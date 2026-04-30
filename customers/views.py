from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from send_logs.models import SendLog
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from .models import Customer
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from django.db import models  # 添加这一行
from .models import Customer, CustomerTag, CustomerTagRelation, CompanySeal
from .models import User
from .models import CompanySetting
from .models import UserEmailConfig
from .models import BusinessRule
from datetime import datetime
from django.conf import settings
from django.http import FileResponse, HttpResponse
from .models import BackupRecord
from .models import CommunicationLog
from .services.search_service import SearchService
from .services.crawler_service import CrawlerService
from .utils.extractors import extract_emails, extract_phones, extract_company_name
from django.views.decorators.http import require_http_methods
from .services.ai_service import AIService
from .utils.duplicate_checker import DuplicateChecker
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from .services.ai_scoring import AIScoringService
from .services.pdf_generator import PDFGenerator
from django.http import FileResponse
from .models import CompanySetting
from django.http import HttpResponse, Http404
from django.utils import timezone
from django.core.mail import EmailMultiAlternatives
from django.urls import reverse
from send_logs.models import SendLog  # 如果还没有导入


import hashlib
import re
import pandas as pd
import os
import json
import requests
import zipfile
import openpyxl

# ========== 在这里添加操作日志相关代码 ==========
from .models import OperationLog

def log_operation(action, module, target='', details=''):
    """记录操作日志的装饰器"""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            response = view_func(request, *args, **kwargs)
            try:
                OperationLog.objects.create(
                    user=request.user.username if request.user.is_authenticated else '匿名',
                    action=action,
                    module=module,
                    target=target,
                    details=details,
                    ip_address=request.META.get('REMOTE_ADDR'),
                )
            except:
                pass
            return response
        return wrapper
    return decorator
# ========== 操作日志代码结束 ==========

def customer_list(request):
    """客户列表页面"""
    from .models import CustomerTag, Order
    from django.db.models import Sum, Q
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    
    level = request.GET.get('level', '')
    keyword = request.GET.get('keyword', '')
    sort_by = request.GET.get('sort', '-last_contact_time')
    tag_id = request.GET.get('tag', '')
    
    queryset = Customer.objects.filter(is_deleted=False)
    
    # ========== 权限过滤（新增） ==========
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        role = request.user.profile.role
        if role == 'sales':
            # 销售人员：只能看自己分配的客户
            queryset = queryset.filter(assigned_sales=request.user)
        elif role == 'dept_leader':
            # 部门主管：只能看本部门的客户
            dept = request.user.profile.department
            if dept:
                queryset = queryset.filter(department=dept)
        elif role == 'readonly':
            # 只读用户：只能看本部门的客户
            dept = request.user.profile.department
            if dept:
                queryset = queryset.filter(department=dept)
        # 管理员：看到所有客户，不需要过滤
    # ========== 权限过滤结束 ==========
    
    if level:
        queryset = queryset.filter(level=level)
    
    if keyword:
        queryset = queryset.filter(
            Q(company_name__icontains=keyword) |
            Q(contact_person__icontains=keyword) |
            Q(email__icontains=keyword)
        )
    
    # 按标签筛选
    if tag_id:
        queryset = queryset.filter(tag_relations__tag_id=tag_id)
    
    queryset = queryset.order_by(sort_by)
    
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # ========== 为每个客户计算订单价值评分 ==========
    for customer in page_obj:
        orders = Order.objects.filter(customer=customer)
        order_count = orders.count()
        
        if order_count > 0:
            # 使用 total_cost 字段计算总金额
            total_amount = orders.aggregate(total=Sum('total_cost'))['total'] or 0
            
            # 基于总金额计算评分 (0-100)
            if total_amount >= 100000:
                customer.order_score = 90
            elif total_amount >= 50000:
                customer.order_score = 70
            elif total_amount >= 10000:
                customer.order_score = 50
            elif total_amount >= 5000:
                customer.order_score = 30
            elif total_amount >= 1000:
                customer.order_score = 20
            else:
                customer.order_score = 10
            
            # 如果有订单但没有 AI 评分，给一个基础分
            if customer.ai_score == 0 or customer.ai_score is None:
                customer.ai_score = 20
        else:
            customer.order_score = 0
        
        # 确保 ai_score 有默认值
        if customer.ai_score is None:
            customer.ai_score = 0
    # ========== 计算结束 ==========
    
    # ========== 获取销售人员列表（用于添加客户弹窗） ==========
    sales_users = User.objects.filter(
        Q(profile__role='sales') | Q(profile__role='dept_leader') | Q(is_superuser=True)
    ).select_related('profile__department').distinct()
    # ========== 获取结束 ==========
    
    level_choices = Customer.LEVEL_CHOICES
    all_tags = CustomerTag.objects.all()
    
    return render(request, 'customers/list.html', {
        'page_obj': page_obj,
        'level_choices': level_choices,
        'current_level': level,
        'keyword': keyword,
        'sort_by': sort_by,
        'current_tag': tag_id,
        'all_tags': all_tags,
        'sales_users': sales_users,  # 添加这一行
    })
# ==================== 二期：搜索任务管理 ====================

from .models import SearchTask, SearchResult

def search_task_create(request):
    """创建搜索任务"""
    if request.method == 'POST':
        business_type = request.POST.get('business_type')
        name = request.POST.get('name')
        keywords = request.POST.get('keywords')
        exclude_words = request.POST.get('exclude_words', '')
        max_results = int(request.POST.get('max_results', 50))
        custom_site = request.POST.get('custom_site', '')
        domain_suffix = request.POST.get('domain_suffix', '')
        search_engine = request.POST.get('search_engine', 'google')  # ← 新增这行
        
        # 外贸/内贸不同字段
        target_country = request.POST.get('target_country', '') if business_type == 'international' else ''
        target_province = request.POST.get('target_province', '') if business_type == 'domestic' else ''
        target_city = request.POST.get('target_city', '') if business_type == 'domestic' else ''
        
        # 定时任务
        schedule_enabled = request.POST.get('schedule_enabled') == 'on'
        schedule_frequency = request.POST.get('schedule_frequency', '')
        schedule_time = request.POST.get('schedule_time', '')
        
        task = SearchTask.objects.create(
            business_type=business_type,
            name=name,
            keywords=keywords,
            exclude_words=exclude_words,
            max_results=max_results,
            custom_site=custom_site,
            domain_suffix=domain_suffix,
            target_country=target_country,
            target_province=target_province,
            target_city=target_city,
            schedule_enabled=schedule_enabled,
            schedule_frequency=schedule_frequency,
            schedule_time=schedule_time,
            search_engine=search_engine,  # ← 新增这行
        )
        
        messages.success(request, f'搜索任务 "{name}" 创建成功')
        return redirect('search_task_list')
    
    return render(request, 'customers/search_task_form.html')


def search_task_detail(request, task_id):
    """查看搜索任务详情和结果"""
    task = get_object_or_404(SearchTask, id=task_id)
    results = SearchResult.objects.filter(task=task).order_by('-created_at')
    return render(request, 'customers/search_task_detail.html', {'task': task, 'results': results})


def search_task_delete(request, task_id):
    """删除搜索任务"""
    task = get_object_or_404(SearchTask, id=task_id)
    task_name = task.name
    task.delete()
    
    # 判断是否为 AJAX 请求
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': f'任务 "{task_name}" 已删除'})
    
    messages.success(request, f'任务 "{task_name}" 已删除')
    return redirect('search_task_list')

def search_task_run(request, task_id):
    """执行搜索任务 - 支持国家限定"""
    task = get_object_or_404(SearchTask, id=task_id)
    
    task.status = 'running'
    task.save()
    
    SearchResult.objects.filter(task=task).delete()
    
    # 解析关键词
    keywords_list = [k.strip() for k in task.keywords.split('\n') if k.strip()]
    main_keyword = keywords_list[0] if keywords_list else ""
    
    # 获取目标国家/省份
    target_country = None
    if task.business_type == 'international':
        target_country = task.target_country  # 外贸使用国家
    else:
        target_country = task.target_province  # 内贸使用省份
    
    # ========== 根据业务类型执行不同的搜索 ==========
    if task.business_type == 'domestic':
        # 内贸 = 百度搜索
        from .services.baidu_search import BaiduSearchService
        baidu = BaiduSearchService()
        results_data = baidu.search(
            keyword=main_keyword,
            max_results=task.max_results,
            province=target_country,
            city=task.target_city
        )
        # 格式化百度搜索结果
        results = [{'url': r['url'], 'title': r['title']} for r in results_data]
    else:
        # 外贸 = Google搜索
        from .services.search_service import SearchService
        search_service = SearchService(business_type=task.business_type)
        results = search_service.search(
            keyword=main_keyword,
            max_results=task.max_results,
            exclude_words=task.exclude_words.split('\n') if task.exclude_words else [],
            domain_suffix=task.domain_suffix,
            target_country=target_country
        )
    # ========== 搜索结束 ==========
    
    # 保存结果
    for item in results:
        SearchResult.objects.create(
            task=task,
            url=item['url'],
            title=item['title'],
        )
    
    task.status = 'completed'
    task.total_found = len(results)
    task.save()
    
    messages.success(request, f'搜索完成（{"百度" if task.business_type == "domestic" else "Google"}），找到 {task.total_found} 条结果')
    return redirect('search_task_detail', task_id=task.id)

def search_result_import(request, result_id):
    """导入搜索结果到客户库（集成重复检测）"""
    from .services.crawler_service import CrawlerService
    from .utils.extractors import extract_emails, extract_phones, extract_company_name
    from .services.ai_service import AIService
    
    result = get_object_or_404(SearchResult, id=result_id)
    
    if result.is_imported:
        return JsonResponse({'success': False, 'message': '该客户已导入'})
    
    # 爬取网站内容
    crawler = CrawlerService()
    crawl_result = crawler.crawl(result.url)
    
    company_name = result.title
    emails = []
    phones = []
    contact_person = ''
    country = ''
    industry = ''
    description = ''
    
    if crawl_result['success']:
        html_content = crawl_result['content']
        
        # 1. 正则提取基础信息
        emails = extract_emails(html_content)
        phones = extract_phones(html_content)
        company_name = extract_company_name(html_content, result.url) or result.title
        
        # 2. AI 提取详细信息
        ai_service = AIService()
        if ai_service.is_available():
            ai_data = ai_service.extract_business_info(html_content, result.url)
            if ai_data:
                contact_person = ai_data.get('contact_person', '') or contact_person
                country = ai_data.get('country', '')
                industry = ai_data.get('industry', '')
                description = ai_data.get('description', '')
                if ai_data.get('company_name'):
                    company_name = ai_data.get('company_name')
        
        # 保存提取的数据到 SearchResult
        result.company_name = company_name
        result.email = emails[0] if emails else ''
        result.phone = phones[0] if phones else ''
        result.extracted_data = {
            'contact_person': contact_person,
            'country': country,
            'industry': industry,
            'description': description
        }
        result.save()
    
    # ========== 重复检测 ==========
    email_value = emails[0] if emails else ''
    check_result = DuplicateChecker.check_duplicate(
        email=email_value if email_value else None,
        company_name=company_name,
        url=result.url
    )
    
    if check_result['is_duplicate']:
        existing = check_result['existing_customer']
        match_type = check_result['match_type']
        
        match_type_display = {
            'email': '邮箱',
            'domain': '域名',
            'company_name': '公司名称',
            'company_name_partial': '公司名称（部分匹配）'
        }.get(match_type, match_type)
        
        return JsonResponse({
            'success': False,
            'is_duplicate': True,
            'message': f'客户已存在！匹配方式：{match_type_display}',
            'existing_customer': {
                'id': existing.id,
                'company_name': existing.company_name,
                'email': existing.email,
                'phone': existing.phone
            }
        })
    # ========== 重复检测结束 ==========
    
    # 处理邮箱唯一约束
    if not email_value:
        email_value = None
    else:
        if Customer.objects.filter(email=email_value).exists():
            email_value = None
    
    # 构建备注信息
    notes_parts = []
    if industry:
        notes_parts.append(f"行业: {industry}")
    if description:
        notes_parts.append(f"简介: {description}")
    notes = "\n".join(notes_parts) if notes_parts else ""
    
    # 创建客户
    customer = Customer.objects.create(
        company_name=company_name or '待完善',
        contact_person=contact_person,
        email=email_value,
        phone=phones[0] if phones else '',
        country=country,
        notes=notes,
        business_type=result.task.business_type,
        source='search_import',
        created_at=timezone.now(),
    )
    
    result.is_imported = True
    result.imported_to = customer.id
    result.save()
    
    # 更新任务统计
    task = result.task
    task.total_imported = SearchResult.objects.filter(task=task, is_imported=True).count()
    task.save()
    
    return JsonResponse({'success': True, 'customer_id': customer.id})

def search_result_ignore(request, result_id):
    """忽略搜索结果"""
    result = get_object_or_404(SearchResult, id=result_id)
    result.is_ignored = True
    result.save()
    return JsonResponse({'success': True})

@log_operation('import', 'customer', '导入客户')
def import_customers(request):
    """导入客户"""
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, '请选择文件')
            return redirect('import_customers')
        
        try:
            df = pd.read_excel(excel_file, engine='openpyxl')
            
            # 添加调试：打印文件信息
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"文件行数: {len(df)}")
            logger.error(f"列名: {list(df.columns)}")
            
            success_count = 0
            fail_count = 0
            duplicate_count = 0
            
            for index, row in df.iterrows():
                email = row.get('邮箱')
                
                if pd.isna(email) or str(email).strip() in ['', '—', '-']:
                    email = None
                else:
                    email = str(email).strip()
                
                company_name = str(row.get('公司名', '')).strip() if pd.notna(row.get('公司名')) else ''
                
                # 打印每一行
                logger.error(f"第{index+2}行: 公司={company_name}, 邮箱={email}")
                
                # 检查重复
                if email and Customer.objects.filter(email=email).exists():
                    duplicate_count += 1
                    logger.error(f"  跳过: 邮箱重复")
                    continue
                
                if not email and company_name:
                    if Customer.objects.filter(company_name=company_name, email__isnull=True).exists():
                        duplicate_count += 1
                        logger.error(f"  跳过: 公司名重复")
                        continue
                
                try:
                    customer = Customer.objects.create(
                        company_name=company_name,
                        contact_person=str(row.get('联系人', '')).strip() if pd.notna(row.get('联系人')) else '',
                        email=email,
                        phone=str(row.get('电话', '')).strip() if pd.notna(row.get('电话')) else '',
                        country=str(row.get('国家', '')).strip() if pd.notna(row.get('国家')) else '',
                        address=str(row.get('地址', '')).strip() if pd.notna(row.get('地址')) else '',
                        source='excel_import',
                        is_deleted=False,  # 👈 添加这一行
                        assigned_sales=request.user,  # 👈 添加这一行：设置当前用户为负责人
                    )
                    success_count += 1
                    logger.error(f"  成功创建: ID={customer.id}")
                except Exception as e:
                    fail_count += 1
                    logger.error(f"  创建失败: {str(e)}")
            
            logger.error(f"导入完成: 成功{success_count}, 重复{duplicate_count}, 失败{fail_count}")
            messages.success(request, f'导入完成！成功: {success_count} 条，重复: {duplicate_count} 条，失败: {fail_count} 条')
            
        except Exception as e:
            import traceback
            logger.error(f"整体错误: {traceback.format_exc()}")
            messages.error(request, f'导入失败: {str(e)}')
        
        return redirect('customer_list')
    
    return render(request, 'customers/import.html')

def update_customer_level(request, customer_id):
    """更新客户等级"""
    if request.method == 'POST':
        data = json.loads(request.body)
        new_level = data.get('level')
        
        customer = Customer.objects.get(id=customer_id)
        customer.level = new_level
        customer.save()
        
        return JsonResponse({'success': True})


from django.core.mail import send_mail
from django.conf import settings
from django.http import JsonResponse
from templates.models import EmailTemplate
from send_logs.models import SendLog
from django.utils import timezone
import json

@log_operation('send_email', 'email', '发送邮件')
def send_email_to_customer(request, customer_id):
    """发送邮件给客户（带追踪）"""
    if request.method == 'POST':
        data = json.loads(request.body)
        template_id = data.get('template_id')
        sender_config_id = data.get('sender_config_id')
        
        customer = Customer.objects.get(id=customer_id)
        template = EmailTemplate.objects.get(id=template_id)
        
        # 准备变量
        variables = {
            'company_name': customer.company_name,
            'contact_person': customer.contact_person or '先生/女士',
            'country': customer.country or '',
            'my_name': request.user.real_name if hasattr(request.user, 'real_name') else request.user.username,
            'my_company': 'Raffinato',
        }
        
        subject = template.subject.format(**variables)
        content = template.content.format(**variables)
        
        # 创建发送记录
        send_log = SendLog.objects.create(
            customer=customer,
            template=template,
            subject=subject,
            content=content,
            sent_by=request.user,
            sent_at=timezone.now(),
            status='pending'
        )
        
        # 1. 添加追踪像素
        tracking_pixel_url = request.build_absolute_uri(
            reverse('track_open', args=[send_log.id])
        )
        tracking_img = f'<img src="{tracking_pixel_url}" width="1" height="1" style="display:none;">'
        
        # 2. 替换链接为追踪链接
        def replace_link(match):
            original_url = match.group(1)
            if 'track/click' in original_url or 'track/open' in original_url:
                return f'href="{original_url}"'
            
            link_id = hashlib.md5(original_url.encode()).hexdigest()[:16]
            tracking_url = request.build_absolute_uri(
                reverse('track_click', args=[send_log.id, link_id])
            )
            tracking_url += f'?url={original_url}'
            return f'href="{tracking_url}"'
        
        final_content = re.sub(r'href="([^"]+)"', replace_link, content)
        final_content += tracking_img
        
        # 获取发件邮箱
        if sender_config_id:
            email_config = UserEmailConfig.objects.get(id=sender_config_id, user=request.user)
            from_email = email_config.email
        else:
            email_config = UserEmailConfig.objects.filter(user=request.user, is_active=True).first()
            from_email = email_config.email if email_config else settings.DEFAULT_FROM_EMAIL
        
        try:
            msg = EmailMultiAlternatives(
                subject=subject,
                body=final_content,
                from_email=from_email,
                to=[customer.email],
            )
            msg.attach_alternative(final_content, "text/html")
            msg.send()
            
            send_log.status = 'sent'
            send_log.save()
            
            # 沟通记录
            CommunicationLog.objects.create(
                customer=customer,
                channel='email',
                direction='outgoing',
                content=f"【邮件发送】\n主题：{subject}\n收件人：{customer.email}\n内容摘要：{content[:300]}",
                communication_time=timezone.now(),
                created_by=request.user.username,
            )
            
            customer.last_contact_time = timezone.now()
            customer.save()
            
            return JsonResponse({'success': True, 'message': '邮件发送成功'})
            
        except Exception as e:
            send_log.status = 'failed'
            send_log.error_message = str(e)
            send_log.save()
            return JsonResponse({'success': False, 'message': str(e)})

def dashboard(request):
    """仪表盘首页"""
    from .models import Order
    from django.db.models import Sum, Count
    from datetime import datetime, timedelta
    
    # 获取当前用户角色
    user_role = getattr(request.user, 'role', 'user') if request.user.is_authenticated else 'user'
    user_department = getattr(request.user, 'department', '') if request.user.is_authenticated else ''
    
    # 判断是否有权限查看订单数据（所有登录用户都可以看）
    can_view_orders = True  # 临时强制显示
    
    # 统计数据
    total_customers = Customer.objects.filter(is_deleted=False).count()
    vip_count = Customer.objects.filter(is_deleted=False, level='vip').count()
    advanced_count = Customer.objects.filter(is_deleted=False, level='advanced').count()
    intermediate_count = Customer.objects.filter(is_deleted=False, level='intermediate').count()
    potential_count = Customer.objects.filter(is_deleted=False, level='potential').count()
    
    # 邮件统计
    sent_count = SendLog.objects.count()
    replied_count = 0  # 暂时设为0，因为 has_replied 字段不存在
    
    # 最近7天邮件活动
    today = timezone.now()
    last_7_days = []
    email_counts = []
    
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        last_7_days.append(day.strftime('%m/%d'))
        count = SendLog.objects.filter(sent_at__date=day.date()).count()
        email_counts.append(count)
    
    # 最近5条发送记录
    recent_logs = SendLog.objects.select_related('customer', 'template').order_by('-sent_at')[:5]
    
    # ========== 新增图表数据 ==========
    
    # 客户增长趋势（最近30天）
    growth_dates = []
    growth_counts = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        growth_dates.append(day.strftime('%m/%d'))
        count = Customer.objects.filter(created_at__date=day.date(), is_deleted=False).count()
        growth_counts.append(count)
    
    # 邮件回复率统计（暂时设为0）
    replied = 0
    not_replied = sent_count
    
    # 待跟进客户（超过30天未联系的中级及以上客户）
    pending_followup = Customer.objects.filter(
        is_deleted=False,
        level__in=['intermediate', 'advanced', 'vip'],
        last_contact_time__lt=today - timedelta(days=30)
    ).count()
    
    # 沉睡客户（超过180天未联系的潜在客户）
    dormant_count = Customer.objects.filter(is_deleted=False, level='potential', is_dormant=True).count()
    
    # 热门国家TOP5
    top_countries = Customer.objects.filter(is_deleted=False, country__isnull=False).exclude(country='').values('country').annotate(count=Count('id')).order_by('-count')[:5]
    top_countries_labels = [item['country'] for item in top_countries]
    top_countries_data = [item['count'] for item in top_countries]
    
    # 待跟进客户趋势（最近30天）
    pending_dates = []
    pending_counts = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        pending_dates.append(day.strftime('%m/%d'))
        count = Customer.objects.filter(
            is_deleted=False,
            level__in=['intermediate', 'advanced', 'vip'],
            last_contact_time__lt=day - timedelta(days=30)
        ).count()
        pending_counts.append(count)
    
    context = {
        # 客户数据
        'total_customers': total_customers,
        'vip_count': vip_count,
        'advanced_count': advanced_count,
        'intermediate_count': intermediate_count,
        'potential_count': potential_count,
        # 邮件数据
        'sent_count': sent_count,
        'replied_count': replied_count,
        'last_7_days': last_7_days,
        'email_counts': email_counts,
        'recent_logs': recent_logs,
        # 图表数据
        'growth_dates': growth_dates,
        'growth_counts': growth_counts,
        'replied': replied,
        'not_replied': not_replied,
        'pending_followup': pending_followup,
        'dormant_count': dormant_count,
        'top_countries_labels': top_countries_labels,
        'top_countries_data': top_countries_data,
        # 待跟进客户
        'pending_dates': pending_dates,
        'pending_counts': pending_counts,      
    }
    
    return render(request, 'dashboard.html', context)

from django.contrib.auth import logout
from django.shortcuts import redirect

def admin_logout(request):
    logout(request)
    # 强制清除 session
    request.session.flush()
    return redirect('/customers/')

import pandas as pd
from django.http import HttpResponse

def export_customers(request):
    """导出客户数据为Excel"""
    import pandas as pd
    from django.http import HttpResponse
    from django.db.models import Q
    
    # 获取筛选参数（与客户列表相同）
    level = request.GET.get('level', '')
    keyword = request.GET.get('keyword', '')
    
    queryset = Customer.objects.filter(is_deleted=False)
    
    if level:
        queryset = queryset.filter(level=level)
    
    if keyword:
        queryset = queryset.filter(
            Q(company_name__icontains=keyword) |
            Q(contact_person__icontains=keyword) |
            Q(email__icontains=keyword)
        )
    
    # 准备导出数据
    data = []
    for customer in queryset:
        data.append({
            '公司名': customer.company_name,
            '联系人': customer.contact_person or '',
            '邮箱': customer.email,
            '电话': customer.phone or '',
            '国家': customer.country or '',
            '地址': customer.address or '',
            '网站': customer.website or '',
            '客户等级': dict(Customer.LEVEL_CHOICES).get(customer.level, ''),
            '上次联系时间': customer.last_contact_time.strftime('%Y-%m-%d %H:%M') if customer.last_contact_time else '',
            '创建时间': customer.created_at.strftime('%Y-%m-%d %H:%M'),
        })
    
    # 创建 DataFrame
    df = pd.DataFrame(data)
    
    # 导出为 Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
    
    # 使用 BytesIO 避免缓冲区问题
    from io import BytesIO
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='客户列表')
    
    response.write(output.getvalue())
    output.close()
    
    return response

@log_operation('batch_send', 'email', '批量发送邮件')
def batch_send_email(request):
    """批量发送邮件"""
    if request.method == 'POST':
        import json
        from django.core.mail import send_mail
        from django.conf import settings
        from templates.models import EmailTemplate
        from send_logs.models import SendLog
        from django.utils import timezone
        
        data = json.loads(request.body)
        customer_ids = data.get('customer_ids', [])
        template_id = data.get('template_id')
        
        if not customer_ids or not template_id:
            return JsonResponse({'success': False, 'message': '参数错误'})
        
        template = EmailTemplate.objects.get(id=template_id)
        customers = Customer.objects.filter(id__in=customer_ids, is_deleted=False)
        
        success_count = 0
        fail_count = 0
        
        for customer in customers:
            try:
                # 准备变量
                variables = {
                    'company_name': customer.company_name,
                    'contact_person': customer.contact_person or '先生/女士',
                    'country': customer.country or '',
                    'my_name': request.user.real_name if hasattr(request.user, 'real_name') else request.user.username,
                    'my_company': 'Raffinato',
                }
                
                subject = template.subject.format(**variables)
                content = template.content.format(**variables)
                
                # 发送邮件
                send_mail(
                    subject=subject,
                    message=content,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[customer.email],
                    fail_silently=False,
                    html_message=content,
                )
                
                # 记录日志
                SendLog.objects.create(
                    customer=customer,
                    template=template,
                    subject=subject,
                    content=content,
                )
                
                # 自动生成沟通记录
                from .models import CommunicationLog
                CommunicationLog.objects.create(
                    customer=customer,
                    channel='email',
                    direction='outgoing',
                    content=f"【批量邮件发送】\n主题：{subject}\n内容摘要：{content[:300]}",
                    communication_time=timezone.now(),
                    created_by=request.user.username if hasattr(request, 'user') else '系统',
                )
                
                # 更新联系时间
                customer.last_contact_time = timezone.now()
                customer.save()
                
                success_count += 1
                
            except Exception as e:
                fail_count += 1
                print(f'发送失败 {customer.email}: {str(e)}')
        
        return JsonResponse({
            'success': True,
            'success_count': success_count,
            'fail_count': fail_count
        })

def customer_detail(request, customer_id):
    """客户详情页"""
    from .models import CommunicationLog
    
    customer = get_object_or_404(Customer, id=customer_id, is_deleted=False)
    send_logs = SendLog.objects.filter(customer=customer).order_by('-sent_at')
    communications = CommunicationLog.objects.filter(customer=customer).order_by('-created_at')
    
    context = {
        'customer': customer,
        'send_logs': send_logs,
        'communications': communications,
    }
    return render(request, 'customers/detail.html', {'customer': customer})


@login_required
def add_communication(request, customer_id):
    """添加沟通记录"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        try:
            # 处理通话时长
            call_duration = None
            call_minutes = request.POST.get('call_minutes')
            call_seconds = request.POST.get('call_seconds')
            if call_minutes or call_seconds:
                call_duration = int(call_minutes or 0) * 60 + int(call_seconds or 0)
            
            # 处理跟进日期
            followup_date = None
            followup_date_str = request.POST.get('followup_date')
            if followup_date_str:
                from datetime import datetime
                followup_date = datetime.strptime(followup_date_str, '%Y-%m-%d').date()
            
            # 创建沟通记录（只使用模型中存在的字段）
            communication = CommunicationLog.objects.create(
                customer=customer,
                channel=request.POST.get('channel'),
                direction=request.POST.get('direction'),
                subject=request.POST.get('subject', ''),
                content=request.POST.get('content', ''),
                call_duration=call_duration,
                attachment=request.FILES.get('attachment'),
                followup_needed=request.POST.get('followup_needed') == 'on',
                followup_date=followup_date,
                followup_note=request.POST.get('followup_note', ''),
                created_by=request.user,
            )
            
            return JsonResponse({'success': True, 'message': '添加成功'})
        
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': '请求方法错误'})

# ==================== 邮箱验证 ====================

def verify_email(request, customer_id):
    """验证客户邮箱"""
    import re
    import socket
    from django.utils import timezone
    
    customer = get_object_or_404(Customer, id=customer_id)
    
    if not customer.email:
        return JsonResponse({'success': False, 'message': '该客户没有邮箱'})
    
    email = customer.email
    
    # 1. 语法验证
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(pattern, email):
        customer.email_verify_status = 'invalid'
        customer.email_verify_date = timezone.now()
        customer.save()
        return JsonResponse({'success': True, 'status': 'invalid', 'message': '邮箱格式不正确'})
    
    # 2. 域名验证
    try:
        domain = email.split('@')[1]
        socket.gethostbyname(domain)
        customer.email_verify_status = 'valid'
        customer.email_verified = True
        message = '邮箱有效'
    except:
        customer.email_verify_status = 'invalid'
        message = '域名无法解析'
    
    customer.email_verify_date = timezone.now()
    customer.save()
    
    return JsonResponse({'success': True, 'status': customer.email_verify_status, 'message': message})


def batch_verify_emails(request):
    """批量验证邮箱"""
    import socket
    from django.utils import timezone
    
    # 获取所有待验证且有邮箱的客户
    customers = Customer.objects.filter(
        is_deleted=False
    ).exclude(email='').exclude(email__isnull=True)[:50]  # 最多50个
    
    valid_count = 0
    invalid_count = 0
    
    for customer in customers:
        email = customer.email
        if not email:
            customer.email_verify_status = 'invalid'
            customer.save()
            invalid_count += 1
            continue
        
        # 简单验证：检查域名是否可解析
        try:
            domain = email.split('@')[1]
            socket.gethostbyname(domain)
            customer.email_verify_status = 'valid'
            customer.email_verified = True
            valid_count += 1
        except:
            customer.email_verify_status = 'invalid'
            invalid_count += 1
        
        customer.email_verify_date = timezone.now()
        customer.save()
    
    return JsonResponse({
        'success': True,
        'valid_count': valid_count,
        'invalid_count': invalid_count,
        'total': valid_count + invalid_count
    })

@log_operation('edit', 'customer', '编辑客户')
def customer_edit(request, customer_id):
    """编辑客户信息"""
    customer = get_object_or_404(Customer, id=customer_id, is_deleted=False)
    
    if request.method == 'POST':
        customer.company_name = request.POST.get('company_name')
        customer.contact_person = request.POST.get('contact_person')
        customer.email = request.POST.get('email') or None
        customer.phone = request.POST.get('phone')
        customer.country = request.POST.get('country')
        customer.province = request.POST.get('province')
        customer.city = request.POST.get('city')
        customer.address = request.POST.get('address')
        # 网站字段：如果为空则设为空字符串，否则验证格式
        website = request.POST.get('website', '')
        if website and not website.startswith(('http://', 'https://')):
            website = 'https://' + website
        customer.website = website
        customer.save()
        
        messages.success(request, '客户信息已更新')
        return redirect('customer_detail', customer_id=customer.id)
    
    return render(request, 'customers/customer_edit.html', {'customer': customer})
# ==================== 客户评分 ====================

def calculate_customer_score(customer):
    """计算单个客户评分"""
    score = 0
    
    # 1. 邮箱有效性（20分）
    if customer.email_verify_status == 'valid':
        score += 20
    elif customer.email_verify_status == 'invalid':
        score += 0
    else:
        score += 10  # 待验证给一半分
    
    # 2. 邮件回复率（25分）
    from send_logs.models import SendLog
    send_count = SendLog.objects.filter(customer=customer).count()
    reply_count = SendLog.objects.filter(customer=customer, has_replied=True).count()
    if send_count > 0:
        reply_rate = reply_count / send_count
        score += int(reply_rate * 25)
    
    # 3. 公司信息完整度（15分）
    info_score = 0
    if customer.company_name:
        info_score += 3
    if customer.phone:
        info_score += 3
    if customer.country or customer.province:
        info_score += 3
    if customer.address:
        info_score += 3
    if customer.website:
        info_score += 3
    score += info_score
    
    # 4. 近期活跃度（20分）
    from django.utils import timezone
    from datetime import timedelta
    if customer.last_contact_time:
        days_since_contact = (timezone.now() - customer.last_contact_time).days
        if days_since_contact <= 7:
            score += 20
        elif days_since_contact <= 30:
            score += 15
        elif days_since_contact <= 90:
            score += 10
        else:
            score += 5
    else:
        score += 0  # 从未联系
    
    # 5. 客户等级（20分）
    level_scores = {
        'vip': 20,
        'advanced': 15,
        'intermediate': 10,
        'potential': 5,
    }
    score += level_scores.get(customer.level, 0)
    
    # 确保不超过100分
    return min(score, 100)


def update_customer_score(request, customer_id):
    """更新单个客户评分"""
    customer = get_object_or_404(Customer, id=customer_id)
    new_score = calculate_customer_score(customer)
    customer.score = new_score
    customer.save()
    
    return JsonResponse({'success': True, 'score': new_score})


@login_required
def batch_update_scores(request):
    """批量更新客户评分"""
    from .services.ai_scoring import AIScoringService
    
    if request.method == 'GET':
        try:
            scoring_service = AIScoringService()
            customers = Customer.objects.filter(is_deleted=False)
            updated_count = 0
            
            for customer in customers:
                try:
                    result = scoring_service.calculate_score(customer)
                    customer.ai_score = result['score']
                    customer.save()
                    updated_count += 1
                except Exception as e:
                    print(f"更新客户 {customer.id} 失败: {e}")
            
            return JsonResponse({
                'success': True,
                'updated_count': updated_count,
                'message': f'成功更新 {updated_count} 个客户'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': '请求方法错误'})

# ==================== 三期：公司文档管理 ====================

from .models import CompanyDocument

def document_list(request):
    """文档列表"""
    doc_type = request.GET.get('type', '')
    documents = CompanyDocument.objects.all()
    
    if doc_type:
        documents = documents.filter(doc_type=doc_type)
    
    doc_types = CompanyDocument.DOCUMENT_TYPES
    
    return render(request, 'customers/document_list.html', {
        'documents': documents,
        'doc_types': doc_types,
        'current_type': doc_type,
    })


def document_upload(request):
    """上传文档"""
    if request.method == 'POST':
        title = request.POST.get('title')
        doc_type = request.POST.get('doc_type')
        file = request.FILES.get('file')
        version = request.POST.get('version', '')
        description = request.POST.get('description', '')
        
        if not title or not file:
            messages.error(request, '请填写标题并选择文件')
            return redirect('document_upload')
        
        CompanyDocument.objects.create(
            title=title,
            doc_type=doc_type,
            file=file,
            version=version,
            description=description,
            uploaded_by=request.user.username if hasattr(request, 'user') and request.user.is_authenticated else '管理员',
        )
        
        messages.success(request, f'文档 "{title}" 上传成功')
        return redirect('document_list')
    
    doc_types = CompanyDocument.DOCUMENT_TYPES
    return render(request, 'customers/document_upload.html', {
        'doc_types': doc_types,
    })


def document_delete(request, doc_id):
    """删除文档"""
    doc = get_object_or_404(CompanyDocument, id=doc_id)
    title = doc.title
    doc.delete()
    messages.success(request, f'文档 "{title}" 已删除')
    return redirect('document_list')


def document_download(request, doc_id):
    """下载文档"""
    from django.http import FileResponse
    doc = get_object_or_404(CompanyDocument, id=doc_id)
    doc.download_count += 1
    doc.save()
    
    response = FileResponse(doc.file, as_attachment=True)
    response['Content-Disposition'] = f'attachment; filename="{doc.file.name.split("/")[-1]}"'
    return response

# ==================== 三期：订单管理 ====================

from .models import Order

def order_list(request):
    """订单列表"""
    from django.db.models import Sum, Q
    
    business_type = request.GET.get('type', '')
    status = request.GET.get('status', '')
    
    orders = Order.objects.select_related('customer').all().order_by('-created_at')
    
    # ========== 权限过滤 ==========
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        role = request.user.profile.role
        if role == 'sales':
            # 销售人员：只能看到自己客户的订单
            orders = orders.filter(customer__assigned_sales=request.user)
        elif role == 'dept_leader':
            # 部门主管：只能看到本部门客户的订单
            dept = request.user.profile.department
            if dept:
                orders = orders.filter(customer__department=dept)
        elif role == 'readonly':
            # 只读用户：只能看到本部门客户的订单
            dept = request.user.profile.department
            if dept:
                orders = orders.filter(customer__department=dept)
        # 管理员：看到所有订单，不需要过滤
    # ========== 权限过滤结束 ==========
    
    if business_type:
        orders = orders.filter(business_type=business_type)
    if status:
        orders = orders.filter(status=status)
    
    # 分开计算外贸和内贸的合计金额
    total_amount_domestic = orders.filter(business_type='domestic').aggregate(total=Sum('subtotal'))['total'] or 0
    total_amount_international = orders.filter(business_type='international').aggregate(total=Sum('subtotal'))['total'] or 0
    
    return render(request, 'customers/order_list.html', {
        'orders': orders,
        'total_amount_domestic': total_amount_domestic,
        'total_amount_international': total_amount_international,
        'current_type': business_type,
        'current_status': status,
    })

import datetime

@log_operation('create', 'order', '创建订单')
@log_operation('create', 'order', '创建订单')
def order_create(request):
    """创建订单（自动生成订单号）"""
    if request.method == 'POST':
        # 获取表单数据
        customer_id = request.POST.get('customer')
        
        # 验证客户是否已选择
        if not customer_id:
            messages.error(request, '请选择客户')
            return redirect('order_create')
        
        business_type = request.POST.get('business_type')
        order_date = request.POST.get('order_date')        
        # 获取产品明细
        product_names = request.POST.getlist('product_name[]')
        specifications = request.POST.getlist('specification[]')
        quantities = request.POST.getlist('quantity[]')
        units = request.POST.getlist('unit[]')
        unit_prices = request.POST.getlist('unit_price[]')
        
        # 计算产品明细和小计
        items = []
        subtotal = 0
        for i in range(len(product_names)):
            if product_names[i]:
                qty_str = quantities[i] if i < len(quantities) else ''
                price_str = unit_prices[i] if i < len(unit_prices) else ''
                
                qty = float(qty_str) if qty_str else 0
                price = float(price_str) if price_str else 0
                amount = qty * price
                subtotal += amount
                items.append({
                    'product_name': product_names[i],
                    'specification': specifications[i] if i < len(specifications) else '',
                    'quantity': qty,
                    'unit': units[i] if i < len(units) and units[i] else '米',
                    'unit_price': price,
                    'amount': amount,
                })
        
        # 获取客户
        customer = Customer.objects.get(id=customer_id)
        
        # ========== 权限检查（修改这里）==========
        if request.user.is_authenticated and hasattr(request.user, 'profile'):
            role = request.user.profile.role
            dept = request.user.profile.department
            
            # 主管（dept_leader）的权限检查
            if role == 'dept_leader':
                # 主管只能创建自己部门的业务类型订单
                if dept and dept.name == '外贸部':
                    if business_type != 'international':
                        messages.error(request, '您属于外贸部，只能创建外贸订单')
                        return redirect('order_create')
                elif dept and dept.name == '内贸部':
                    if business_type != 'domestic':
                        messages.error(request, '您属于内贸部，只能创建内贸订单')
                        return redirect('order_create')
                
                # 主管只能看到自己部门的客户（已在GET中过滤）
                if customer.department != dept:
                    messages.error(request, '您没有权限为其他部门的客户创建订单')
                    return redirect('order_create')
            
            elif role == 'sales':
                # 销售人员只能为自己分配的客户创建订单
                if customer.assigned_sales != request.user:
                    messages.error(request, '您没有权限为这个客户创建订单')
                    return redirect('order_create')
                
                # 销售人员只能创建自己部门的业务类型订单
                if dept and dept.name == '外贸部':
                    if business_type != 'international':
                        messages.error(request, '您属于外贸部，只能创建外贸订单')
                        return redirect('order_create')
                elif dept and dept.name == '内贸部':
                    if business_type != 'domestic':
                        messages.error(request, '您属于内贸部，只能创建内贸订单')
                        return redirect('order_create')
        # ========== 权限检查结束 ==========
        
        # 自动生成订单号：SO#20260412A01
        import datetime
        today = datetime.date.today()
        today_str = today.strftime('%Y%m%d')
        type_prefix = 'A' if business_type == 'international' else 'B'
        last_order = Order.objects.filter(
            order_no__startswith=f'SO#{today_str}{type_prefix}'  # PO# 改为 SO#
        ).order_by('-order_no').first()

        if last_order:
            last_num = int(last_order.order_no[-2:])
            new_num = last_num + 1
        else:
            new_num = 1

        order_no = f'SO#{today_str}{type_prefix}{new_num:02d}'  # PO# 改为 SO#
        
        # 创建订单（自动使用当前用户的部门）
        order = Order.objects.create(
            order_no=order_no,
            customer=customer,
            business_type=business_type,
            order_date=order_date,
            items=items,
            subtotal=subtotal,
            notes=request.POST.get('notes', ''),
            sales_person=request.user.username,
            # 如果有部门字段，添加这行
            # department=request.user.profile.department,
        )
        
        messages.success(request, f'订单 {order_no} 创建成功')
        return redirect('order_detail', order_id=order.id)
    
    # GET 请求
    import datetime
    customers = Customer.objects.filter(is_deleted=False)
    
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        role = request.user.profile.role
        dept = request.user.profile.department
        
        if role == 'sales':
            # 销售人员只能看到自己分配的客户
            customers = customers.filter(assigned_sales=request.user)
        elif role == 'dept_leader':
            # 主管只能看到自己部门的客户
            if dept:
                customers = customers.filter(department=dept)
        # admin 可以看到所有客户
    
    return render(request, 'customers/order_form.html', {
        'customers': customers,
        'today_date': datetime.date.today().strftime('%Y-%m-%d'),
        'user_role': request.user.profile.role if hasattr(request.user, 'profile') else 'admin',
        'user_dept': request.user.profile.department.name if hasattr(request.user, 'profile') and request.user.profile.department else '',
        'user_department': request.user.profile.department,  # 添加这行，传递部门对象到模板
    })


def order_detail(request, order_id):
    """订单详情"""
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'customers/order_detail.html', {'order': order})

@log_operation('edit', 'order', '编辑订单')
def order_edit(request, order_id):
    """编辑订单状态和备注"""
    order = get_object_or_404(Order, id=order_id)
    
    # 权限检查：销售人员只能编辑自己客户的订单
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        role = request.user.profile.role
        if role == 'sales':
            if order.customer.assigned_sales != request.user:
                messages.error(request, '您没有权限编辑此订单')
                return redirect('order_list')
        elif role == 'dept_leader':
            # 主管可以编辑本部门订单
            dept = request.user.profile.department
            if dept and order.customer.department != dept:
                messages.error(request, '您没有权限编辑此订单')
                return redirect('order_list')

    if request.method == 'POST':
        order.status = request.POST.get('status')
        order.notes = request.POST.get('notes', '')
        order.save()
        messages.success(request, f'订单 {order.order_no} 已更新')
        return redirect('order_detail', order_id=order.id)
    
    # GET 请求：显示编辑表单
    return render(request, 'customers/order_edit.html', {'order': order})

@log_operation('delete', 'order', '删除订单')
def order_delete(request, order_id):
    """删除订单"""
    order = get_object_or_404(Order, id=order_id)
    order_no = order.order_no
    order.delete()
    messages.success(request, f'订单 {order_no} 已删除')
    return redirect('order_list')


def order_print(request, order_id):
    from .models import CompanySeal
    
    order = get_object_or_404(Order, id=order_id)
    show_seal = request.GET.get('seal', '0') == '1'
    seal_path = request.GET.get('seal_path', '')
    
    # 不再从数据库查询业务章
    # seals = CompanySeal.objects.all()
    
    # 不再需要 selected_seal
    # selected_seal = None
    
    # 计算大写金额
    chinese_amount = number_to_chinese(order.subtotal)
    
    return render(request, 'customers/order_print.html', {
        'order': order,
        'show_seal': show_seal,
        'seal_path': seal_path,  # 直接传图片路径
        'chinese_amount': chinese_amount,
    })

def number_to_chinese(amount):
    """将数字转换为中文大写金额"""
    if not amount or amount == 0:
        return '零元整'
    
    chinese_num = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    chinese_unit = ['', '拾', '佰', '仟']
    chinese_big_unit = ['', '万', '亿']
    
    # 分离整数和小数部分
    amount_str = f"{float(amount):.2f}"
    integer_part = amount_str.split('.')[0]
    decimal_part = amount_str.split('.')[1]
    
    # 转换整数部分
    def convert_integer(num_str):
        if not num_str or int(num_str) == 0:
            return ''
        
        length = len(num_str)
        result = ''
        zero_flag = False
        
        for i, char in enumerate(num_str):
            digit = int(char)
            if digit == 0:
                zero_flag = True
            else:
                if zero_flag:
                    result += '零'
                    zero_flag = False
                result += chinese_num[digit] + chinese_unit[length - i - 1]
        
        return result
    
    # 分段处理
    integer_len = len(integer_part)
    if integer_len <= 4:
        integer_chinese = convert_integer(integer_part)
    else:
        parts = []
        temp = integer_part
        while temp:
            parts.insert(0, temp[-4:])
            temp = temp[:-4]
        
        result_parts = []
        for i, part in enumerate(parts):
            part_chinese = convert_integer(part)
            if part_chinese:
                unit_index = len(parts) - i - 1
                result_parts.append(part_chinese + chinese_big_unit[unit_index])
        
        integer_chinese = ''.join(result_parts)
    
    # 处理小数部分
    decimal_chinese = ''
    if int(decimal_part[0]) > 0:
        decimal_chinese += chinese_num[int(decimal_part[0])] + '角'
    if int(decimal_part[1]) > 0:
        decimal_chinese += chinese_num[int(decimal_part[1])] + '分'
    
    if integer_chinese and decimal_chinese:
        return f"{integer_chinese}元{decimal_chinese}"
    elif integer_chinese:
        return f"{integer_chinese}元整"
    elif decimal_chinese:
        return f"{decimal_chinese}"
    else:
        return '零元整'

@log_operation('delete', 'customer', '删除客户')
def customer_delete(request, customer_id):
    """删除单个客户（软删除）- 支持 AJAX 和普通请求"""
    customer = get_object_or_404(Customer, id=customer_id)
    customer_name = customer.company_name
    customer.is_deleted = True
    customer.save()
    
    # 判断是否为 AJAX 请求
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': f'客户 "{customer_name}" 已删除'})
    
    # 普通表单请求
    messages.success(request, f'客户 "{customer_name}" 已删除')
    return redirect('customer_list')

@log_operation('batch_delete', 'customer', '批量删除客户')
def batch_delete_customers(request):
    """批量删除客户"""
    if request.method == 'POST':
        data = json.loads(request.body)
        customer_ids = data.get('customer_ids', [])
        
        if not customer_ids:
            return JsonResponse({'success': False, 'message': '请选择客户'})
        
        deleted_count = Customer.objects.filter(id__in=customer_ids).update(is_deleted=True)
        
        return JsonResponse({'success': True, 'deleted_count': deleted_count})

@log_operation('batch_delete', 'order', '批量删除订单')
def batch_delete_orders(request):
    """批量删除订单"""
    if request.method == 'POST':
        data = json.loads(request.body)
        order_ids = data.get('order_ids', [])
        
        if not order_ids:
            return JsonResponse({'success': False, 'message': '请选择订单'})
        
        deleted_count = Order.objects.filter(id__in=order_ids).delete()[0]
        
        return JsonResponse({'success': True, 'deleted_count': deleted_count})

# ==================== 订单汇总报表 ====================

def order_summary(request):
    """订单汇总报表"""
    from datetime import datetime
    
    report_type = request.GET.get('type', 'customer')
    
    # 获取筛选参数
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    business_type = request.GET.get('business_type', '')
    customer_id = request.GET.get('customer_id')
    product_name = request.GET.get('product_name')
    year = request.GET.get('year', datetime.now().year)
    month = request.GET.get('month', '')
    view_type = request.GET.get('view', 'date')
    rank_limit = int(request.GET.get('rank_limit', 20))
    
    # 基础查询
    orders = Order.objects.all()
    
    # ========== 权限过滤（添加这部分）==========
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        role = request.user.profile.role
        dept = request.user.profile.department
        
        if role == 'sales':
            # 销售人员：只能看到自己的订单
            orders = orders.filter(sales_person=request.user.username)
            
        elif role == 'dept_leader':
            # 主管：只能看到自己部门的订单
            if dept and dept.name == '外贸部':
                orders = orders.filter(business_type='international')
                # 如果业务类型筛选器选择了内贸，强制改为外贸
                if business_type == 'domestic':
                    business_type = 'international'
            elif dept and dept.name == '内贸部':
                orders = orders.filter(business_type='domestic')
                # 如果业务类型筛选器选择了外贸，强制改为内贸
                if business_type == 'international':
                    business_type = 'domestic'
        
        # admin 可以看到所有订单，不做过滤
    
    # 如果用户是主管，且没有选择业务类型，自动选择本部门的类型
    if hasattr(request.user, 'profile') and request.user.profile.role == 'dept_leader':
        dept = request.user.profile.department
        if dept and not business_type:
            if dept.name == '外贸部':
                business_type = 'international'
            elif dept.name == '内贸部':
                business_type = 'domestic'
    # ========== 权限过滤结束 ==========
    
    # 日期筛选
    if date_from:
        orders = orders.filter(order_date__gte=date_from)
    if date_to:
        orders = orders.filter(order_date__lte=date_to)
    if year and not date_from:
        if month:
            orders = orders.filter(order_date__year=year, order_date__month=month)
        else:
            orders = orders.filter(order_date__year=year)
    
    # 业务类型筛选
    if business_type:
        orders = orders.filter(business_type=business_type)
    
    # 辅助函数：从订单明细中计算总金额
    def get_order_amount(order):
        total = 0
        for item in order.items:
            total += item.get('amount', 0)
        return total
    
    # 全局合计（从明细计算）
    rmb_total = 0
    usd_total = 0
    for order in orders:
        amount = get_order_amount(order)
        if order.business_type == 'domestic':
            rmb_total += amount
        else:
            usd_total += amount
    total_orders_count = orders.count()
    
    # ========== 客户列表也根据权限过滤 ==========
    # 客户列表
    if hasattr(request.user, 'profile') and request.user.profile.role == 'sales':
        # 销售人员：只能看到自己分配的客户
        customers = Customer.objects.filter(is_deleted=False, assigned_sales=request.user).order_by('company_name')
    elif hasattr(request.user, 'profile') and request.user.profile.role == 'dept_leader':
        # 主管：只能看到自己部门的客户
        dept = request.user.profile.department
        if dept:
            customers = Customer.objects.filter(is_deleted=False, department=dept).order_by('company_name')
        else:
            customers = Customer.objects.filter(is_deleted=False).order_by('company_name')
    else:
        # 管理员：可以看到所有客户
        customers = Customer.objects.filter(is_deleted=False).order_by('company_name')
    
    # 产品列表也根据权限过滤
    products = set()
    for o in orders:
        for item in o.items:
            if item.get('product_name'):
                products.add(item.get('product_name'))
    products = sorted([p for p in products if p and p != 'None'])
    
    # 选中的客户名称
    selected_customer_name = ''
    if customer_id:
        selected_customer = Customer.objects.filter(id=customer_id).first()
        selected_customer_name = selected_customer.company_name if selected_customer else ''
    
    # ... 后面的代码保持不变 ...
    result = []
    summary = {}
    
    # ================= 按客户汇总 =================
    if report_type == 'customer':
        if customer_id:
            # 选择了具体客户：显示该客户的订单明细
            orders = orders.filter(customer_id=customer_id)
            customer = Customer.objects.get(id=customer_id)
            summary['customer_name'] = customer.company_name
            summary['total_orders'] = orders.count()
            
            rmb_sum = 0
            usd_sum = 0
            for order in orders:
                amount = get_order_amount(order)
                if order.business_type == 'domestic':
                    rmb_sum += amount
                else:
                    usd_sum += amount
            summary['rmb_amount'] = rmb_sum
            summary['usd_amount'] = usd_sum
            
            for order in orders:
                for item in order.items:
                    result.append({
                        'order_no': order.order_no,
                        'order_date': order.order_date,
                        'product_name': item.get('product_name', ''),
                        'specification': item.get('specification', '-'),
                        'quantity': item.get('quantity', 0),
                        'unit': item.get('unit', '米'),
                        'unit_price': item.get('unit_price', 0),
                        'amount': item.get('amount', 0),
                        'order_amount': get_order_amount(order),
                        'status': order.get_status_display(),
                    })
        else:
            # 未选择客户：显示所有客户的汇总（每个客户一行）
            customer_data = {}
            for order in orders:
                cid = order.customer.id
                if cid not in customer_data:
                    customer_data[cid] = {
                        'customer_name': order.customer.company_name,
                        'order_count': 0,
                        'rmb_amount': 0,
                        'usd_amount': 0,
                    }
                customer_data[cid]['order_count'] += 1
                amount = get_order_amount(order)
                if order.business_type == 'domestic':
                    customer_data[cid]['rmb_amount'] += amount
                else:
                    customer_data[cid]['usd_amount'] += amount
            
            for data in customer_data.values():
                data['total_amount'] = data['rmb_amount'] + data['usd_amount']
                result.append(data)
            result.sort(key=lambda x: x['customer_name'])
            
            # 汇总信息（所有客户合计）
            summary['total_customers'] = len(customer_data)
            summary['total_orders'] = sum(r['order_count'] for r in result)
            summary['rmb_amount'] = sum(r['rmb_amount'] for r in result)
            summary['usd_amount'] = sum(r['usd_amount'] for r in result)
    
    # ================= 按产品汇总 =================
    elif report_type == 'product':
        product_data = {}
        for order in orders:
            for item in order.items:
                pname = item.get('product_name', '')
                if product_name and pname != product_name:
                    continue
                spec = item.get('specification', '-')
                key = f"{pname}_{spec}"
                if key not in product_data:
                    product_data[key] = {
                        'product_name': pname,
                        'specification': spec,
                        'total_quantity': 0,
                        'rmb_amount': 0,
                        'usd_amount': 0,
                    }
                product_data[key]['total_quantity'] += item.get('quantity', 0)
                amount = item.get('amount', 0)
                if order.business_type == 'domestic':
                    product_data[key]['rmb_amount'] += amount
                else:
                    product_data[key]['usd_amount'] += amount
        
        result = list(product_data.values())
        result.sort(key=lambda x: (x['product_name'], x['specification']))
        
        summary['product_name'] = product_name or '全部产品'
        summary['total_quantity'] = sum(r['total_quantity'] for r in result)
        summary['rmb_amount'] = sum(r['rmb_amount'] for r in result)
        summary['usd_amount'] = sum(r['usd_amount'] for r in result)
    
    # ================= 按时间汇总 =================
    elif report_type == 'date':
        if view_type == 'date':
            date_groups = {}
            for order in orders:
                d = order.order_date.strftime('%Y-%m-%d')
                if d not in date_groups:
                    date_groups[d] = {
                        'date': order.order_date,
                        'daily_count': 0,
                        'rmb_amount': 0,
                        'usd_amount': 0,
                    }
                date_groups[d]['daily_count'] += 1
                amount = get_order_amount(order)
                if order.business_type == 'domestic':
                    date_groups[d]['rmb_amount'] += amount
                else:
                    date_groups[d]['usd_amount'] += amount
            result = list(date_groups.values())
            result.sort(key=lambda x: x['date'], reverse=True)
        else:
            product_groups = {}
            for order in orders:
                for item in order.items:
                    key = f"{item.get('product_name', '')}_{item.get('specification', '-')}"
                    if key not in product_groups:
                        product_groups[key] = {
                            'product_name': item.get('product_name', ''),
                            'specification': item.get('specification', '-'),
                            'total_quantity': 0,
                            'rmb_amount': 0,
                            'usd_amount': 0,
                        }
                    product_groups[key]['total_quantity'] += item.get('quantity', 0)
                    amount = item.get('amount', 0)
                    if order.business_type == 'domestic':
                        product_groups[key]['rmb_amount'] += amount
                    else:
                        product_groups[key]['usd_amount'] += amount
            result = list(product_groups.values())
            result.sort(key=lambda x: x['product_name'])
        
        summary['total_orders'] = orders.count()
        summary['rmb_total'] = rmb_total
        summary['usd_total'] = usd_total
    
    # ================= 按销售额汇总 =================
    elif report_type == 'rank':
        customer_summary = {}
        for order in orders:
            cid = order.customer.id
            if cid not in customer_summary:
                customer_summary[cid] = {
                    'customer_name': order.customer.company_name,
                    'order_count': 0,
                    'rmb_amount': 0,
                    'usd_amount': 0,
                }
            customer_summary[cid]['order_count'] += 1
            amount = get_order_amount(order)
            if order.business_type == 'domestic':
                customer_summary[cid]['rmb_amount'] += amount
            else:
                customer_summary[cid]['usd_amount'] += amount
        
        for data in customer_summary.values():
            data['total_amount'] = data['rmb_amount'] + data['usd_amount']
        
        result = sorted(customer_summary.values(), key=lambda x: x['total_amount'], reverse=True)
        total_all = sum(r['total_amount'] for r in result)
        for r in result:
            r['percentage'] = (r['total_amount'] / total_all * 100) if total_all > 0 else 0
        
        if rank_limit > 0:
            result = result[:rank_limit]
        
        summary['total_customers'] = len(customer_summary)
        summary['rmb_total'] = rmb_total
        summary['usd_total'] = usd_total
        summary['total_amount'] = total_all
    
    # 年份列表（也根据权限过滤）
    years = orders.dates('order_date', 'year')
    
    # 传递用户权限信息到模板
    user_role = ''
    user_dept = ''
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        user_role = request.user.profile.role
        user_dept = request.user.profile.department.name if request.user.profile.department else ''
    
    context = {
        'report_type': report_type,
        'result': result,
        'summary': summary,
        'customers': customers,
        'products': products,
        'selected_customer': customer_id,
        'selected_customer_name': selected_customer_name,
        'selected_product': product_name,
        'years': years,
        'current_year': int(year),
        'current_month': int(month) if month else '',
        'months': range(1, 13),
        'date_from': date_from,
        'date_to': date_to,
        'business_type': business_type,
        'rank_limit': rank_limit,
        'view_type': view_type,
        'rmb_total': rmb_total,
        'usd_total': usd_total,
        'total_orders_count': total_orders_count,
        # 添加权限信息
        'user_role': user_role,
        'user_dept': user_dept,
    }
    
    return render(request, 'customers/order_summary.html', context)

# ==================== 多邮箱配置 ====================

from .models import UserEmailConfig
from django.contrib.auth.decorators import login_required

# ==================== 管理员视图（可管理所有用户）====================

@login_required
def admin_email_config_list(request):
    """管理员：查看所有用户的邮箱配置"""
    # 权限检查
    if not (request.user.is_superuser or request.user.role == 'admin'):
        messages.error(request, '权限不足')
        return redirect('email_config_list')  # 普通用户跳转到自己的列表
    
    configs = UserEmailConfig.objects.select_related('user').all().order_by('-created_at')
    return render(request, 'customers/admin_email_config_list.html', {'configs': configs})


@login_required
def admin_email_config_create(request):
    """管理员：为用户创建邮箱配置"""
    from .models import User
    
    # 权限检查
    if not (request.user.is_superuser or request.user.role == 'admin'):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        is_default = request.POST.get('is_default') == 'on'
        
        try:
            target_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            messages.error(request, '用户不存在')
            return redirect('admin_email_config_create')
        
        # 如果是默认邮箱，取消该用户的其他默认
        if is_default:
            UserEmailConfig.objects.filter(user=target_user, is_default=True).update(is_default=False)
        
        UserEmailConfig.objects.create(
            user=target_user,
            email=request.POST.get('email'),
            smtp_host=request.POST.get('smtp_host'),
            smtp_port=int(request.POST.get('smtp_port')),
            smtp_password=request.POST.get('smtp_password'),
            from_name=request.POST.get('from_name'),
            signature=request.POST.get('signature', ''),
            is_default=is_default,
            daily_limit=int(request.POST.get('daily_limit', 200)),
        )
        messages.success(request, f'已为用户 {target_user.username} 添加邮箱配置')
        return redirect('admin_email_config_list')
    
    # GET 请求：显示表单，带用户列表
    users = User.objects.filter(is_active=True)
    return render(request, 'customers/admin_email_config_form.html', {'users': users, 'config': None})


@login_required
def admin_email_config_edit(request, config_id):
    """管理员：编辑任意用户的邮箱配置"""
    from .models import User
    
    # 权限检查
    if not (request.user.is_superuser or request.user.role == 'admin'):
        messages.error(request, '权限不足')
        return redirect('email_config_list')
    
    config = get_object_or_404(UserEmailConfig, id=config_id)
    
    if request.method == 'POST':
        is_default = request.POST.get('is_default') == 'on'
        
        # 如果是默认邮箱，取消该用户的其他默认
        if is_default:
            UserEmailConfig.objects.filter(user=config.user, is_default=True).exclude(id=config.id).update(is_default=False)
        
        config.email = request.POST.get('email')
        config.smtp_host = request.POST.get('smtp_host')
        config.smtp_port = int(request.POST.get('smtp_port'))
        config.from_name = request.POST.get('from_name')
        config.signature = request.POST.get('signature', '')
        config.is_default = is_default
        config.daily_limit = int(request.POST.get('daily_limit', 200))
        
        # 只有填写了新密码才更新
        new_password = request.POST.get('smtp_password')
        if new_password:
            config.smtp_password = new_password
        
        config.save()
        
        messages.success(request, f'已更新 {config.user.username} 的邮箱配置')
        return redirect('admin_email_config_list')
    
    return render(request, 'customers/admin_email_config_form.html', {'config': config, 'users': None})


@login_required
def admin_email_config_delete(request, config_id):
    """管理员：删除任意用户的邮箱配置"""
    # 权限检查
    if not (request.user.is_superuser or request.user.role == 'admin'):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    config = get_object_or_404(UserEmailConfig, id=config_id)
    config.delete()
    messages.success(request, '邮箱配置已删除')
    return redirect('admin_email_config_list')


@login_required
def admin_email_config_test(request, config_id):
    """管理员：测试任意用户的邮箱配置"""
    # 权限检查
    if not (request.user.is_superuser or request.user.role == 'admin'):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    config = get_object_or_404(UserEmailConfig, id=config_id)
    
    try:
        import smtplib
        from email.message import EmailMessage
        from django.utils import timezone
        
        msg = EmailMessage()
        msg.set_content(f'这是一封测试邮件。\n\n如果您收到此邮件，说明邮箱配置正确。\n\n发送时间：{timezone.now().strftime("%Y-%m-%d %H:%M:%S")}')
        msg['Subject'] = 'Raffinato CRM - 邮箱配置测试'
        msg['From'] = f"{config.from_name} <{config.email}>"
        msg['To'] = config.user.email if config.user.email else config.email  # 发送到用户邮箱或配置邮箱
        
        if config.smtp_port == 465:
            server = smtplib.SMTP_SSL(config.smtp_host, config.smtp_port)
        else:
            server = smtplib.SMTP(config.smtp_host, config.smtp_port)
            server.starttls()
        
        server.login(config.email, config.smtp_password)
        server.send_message(msg)
        server.quit()
        
        return JsonResponse({'success': True, 'message': '测试邮件发送成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ==================== 普通用户视图（只能看到自己的）====================

@login_required
def email_config_list(request):
    """普通用户：查看自己的邮箱配置"""
    configs = UserEmailConfig.objects.filter(user=request.user).order_by('-is_default', '-created_at')
    return render(request, 'customers/email_config_list.html', {'configs': configs})


@login_required
def email_config_create(request):
    """普通用户：创建自己的邮箱配置"""
    if request.method == 'POST':
        is_default = request.POST.get('is_default') == 'on'
        if is_default:
            UserEmailConfig.objects.filter(user=request.user, is_default=True).update(is_default=False)
        
        UserEmailConfig.objects.create(
            user=request.user,
            email=request.POST.get('email'),
            smtp_host=request.POST.get('smtp_host'),
            smtp_port=int(request.POST.get('smtp_port')),
            smtp_password=request.POST.get('smtp_password'),
            from_name=request.POST.get('from_name'),
            signature=request.POST.get('signature', ''),
            is_default=is_default,
            daily_limit=int(request.POST.get('daily_limit', 200)),
        )
        messages.success(request, '邮箱配置添加成功')
        return redirect('email_config_list')
    
    return render(request, 'customers/email_config_form.html')


@login_required
def email_config_edit(request, config_id):
    """普通用户：编辑自己的邮箱配置"""
    config = get_object_or_404(UserEmailConfig, id=config_id, user=request.user)
    
    if request.method == 'POST':
        is_default = request.POST.get('is_default') == 'on'
        if is_default:
            UserEmailConfig.objects.filter(user=request.user, is_default=True).update(is_default=False)
        
        config.email = request.POST.get('email')
        config.smtp_host = request.POST.get('smtp_host')
        config.smtp_port = int(request.POST.get('smtp_port'))
        config.from_name = request.POST.get('from_name')
        config.signature = request.POST.get('signature', '')
        config.is_default = is_default
        config.daily_limit = int(request.POST.get('daily_limit', 200))
        
        new_password = request.POST.get('smtp_password')
        if new_password:
            config.smtp_password = new_password
        
        config.save()
        
        messages.success(request, '邮箱配置更新成功')
        return redirect('email_config_list')
    
    return render(request, 'customers/email_config_form.html', {'config': config})


@login_required
def email_config_delete(request, config_id):
    """普通用户：删除自己的邮箱配置"""
    config = get_object_or_404(UserEmailConfig, id=config_id, user=request.user)
    config.delete()
    messages.success(request, '邮箱配置已删除')
    return redirect('email_config_list')


@login_required
def email_config_test(request, config_id):
    """普通用户：测试自己的邮箱配置"""
    config = get_object_or_404(UserEmailConfig, id=config_id, user=request.user)
    
    try:
        import smtplib
        from email.message import EmailMessage
        from django.utils import timezone
        
        msg = EmailMessage()
        msg.set_content(f'这是一封测试邮件。\n\n如果您收到此邮件，说明邮箱配置正确。\n\n发送时间：{timezone.now().strftime("%Y-%m-%d %H:%M:%S")}')
        msg['Subject'] = 'Raffinato CRM - 邮箱配置测试'
        msg['From'] = f"{config.from_name} <{config.email}>"
        msg['To'] = request.user.email
        
        if config.smtp_port == 465:
            server = smtplib.SMTP_SSL(config.smtp_host, config.smtp_port)
        else:
            server = smtplib.SMTP(config.smtp_host, config.smtp_port)
            server.starttls()
        
        server.login(config.email, config.smtp_password)
        server.send_message(msg)
        server.quit()
        
        return JsonResponse({'success': True, 'message': '测试邮件发送成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# ==================== 操作日志 ====================

@login_required
def operation_log_list(request):
    """操作日志列表"""
    from django.core.paginator import Paginator
    
    logs = OperationLog.objects.all()
    
    # 筛选
    action = request.GET.get('action', '')
    module = request.GET.get('module', '')
    user = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if action:
        logs = logs.filter(action=action)
    if module:
        logs = logs.filter(module=module)
    if user:
        logs = logs.filter(user__icontains=user)
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)
    
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 获取筛选选项
    actions = OperationLog.ACTION_CHOICES
    modules = OperationLog.MODULE_CHOICES
    users = OperationLog.objects.values_list('user', flat=True).distinct()
    
    return render(request, 'customers/operation_log_list.html', {
        'page_obj': page_obj,
        'actions': actions,
        'modules': modules,
        'users': users,
        'current_action': action,
        'current_module': module,
        'current_user': user,
        'date_from': date_from,
        'date_to': date_to,
    })

# ==================== 清除操作日志 ====================

@login_required
def clear_operation_logs(request):
    """手动清除操作日志"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': '只有管理员可以操作'})
    
    if request.method == 'POST':
        days = int(request.GET.get('days', 60))
        from django.utils import timezone
        from datetime import timedelta
        cutoff_date = timezone.now() - timedelta(days=days)
        deleted_count = OperationLog.objects.filter(created_at__lt=cutoff_date).delete()[0]
        
        # 记录本次清除操作
        OperationLog.objects.create(
            user=request.user.username,
            action='delete',
            module='system',
            target=f'删除{days}天前日志',
            details=f'共删除{deleted_count}条记录',
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        
        return JsonResponse({'success': True, 'deleted_count': deleted_count})
    
    return JsonResponse({'success': False, 'message': '请求方法错误'})

# ==================== 业务章管理 ====================

from .models import CompanySeal

@login_required
def seal_list(request):
    """业务章列表（仅管理员可见）"""
    # 权限控制：只有超级管理员或 role='admin' 的用户可以访问
    if not (request.user.is_superuser or request.user.role == 'admin'):
        messages.error(request, '您没有权限访问印章管理')
        return redirect('dashboard')
    
    # 管理员可以查看所有印章（如果需要按部门过滤，可保留条件，但一般管理员看全部即可）
    seals = CompanySeal.objects.all().order_by('-created_at')
    
    return render(request, 'customers/seal_list.html', {'seals': seals})

@login_required
def seal_upload(request):
    """上传业务章"""
    if not (request.user.is_superuser or request.user.role == 'admin'):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    if request.method == 'POST':
        brand_name = request.POST.get('brand_name')
        department = request.POST.get('department')
        seal_image = request.FILES.get('seal_image')
        description = request.POST.get('description', '')
        
        if not brand_name or not seal_image:
            return JsonResponse({'success': False, 'message': '品牌名称和图片不能为空'})
        
        seal = CompanySeal.objects.create(
            brand_name=brand_name,
            department=department,
            seal_image=seal_image,
            description=description
        )
        return JsonResponse({'success': True, 'seal_id': seal.id})
    
    return JsonResponse({'success': False, 'message': '请求方法错误'})

@login_required
def seal_delete(request, seal_id):
    if not (request.user.is_superuser or request.user.role == 'admin'):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    try:
        seal = CompanySeal.objects.get(id=seal_id)
        seal.delete()
        return JsonResponse({'success': True})
    except CompanySeal.DoesNotExist:
        return JsonResponse({'success': False, 'message': '业务章不存在'})

@login_required
def seal_reorder(request):
    """业务章排序（仅管理员）"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    if request.method == 'POST':
        data = json.loads(request.body)
        for item in data.get('orders', []):
            CompanySeal.objects.filter(id=item['id']).update(sort_order=item['order'])
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

from .models import CustomerTag, CustomerTagRelation

@login_required
def get_customer_tags(request, customer_id):
    """获取客户标签"""
    customer = get_object_or_404(Customer, id=customer_id)
    tags = customer.tag_relations.select_related('tag').all()
    data = [{'id': t.tag.id, 'name': t.tag.name, 'color': t.tag.color} for t in tags]
    return JsonResponse(data, safe=False)


@login_required
def add_customer_tag(request, customer_id):
    """给客户添加标签"""
    if request.method == 'POST':
        data = json.loads(request.body)
        tag_id = data.get('tag_id')
        tag_name = data.get('tag_name')
        
        customer = get_object_or_404(Customer, id=customer_id)
        
        if tag_id:
            tag = get_object_or_404(CustomerTag, id=tag_id)
        elif tag_name:
            tag, _ = CustomerTag.objects.get_or_create(name=tag_name)
        else:
            return JsonResponse({'success': False, 'message': '请选择或输入标签'})
        
        CustomerTagRelation.objects.get_or_create(customer=customer, tag=tag)
        
        return JsonResponse({'success': True, 'tag': {'id': tag.id, 'name': tag.name, 'color': tag.color}})


@login_required
def remove_customer_tag(request, customer_id, tag_id):
    """移除客户标签"""
    if request.method == 'POST':
        CustomerTagRelation.objects.filter(customer_id=customer_id, tag_id=tag_id).delete()
        return JsonResponse({'success': True})


@login_required
def batch_add_tags(request):
    """批量添加标签"""
    if request.method == 'POST':
        data = json.loads(request.body)
        customer_ids = data.get('customer_ids', [])
        tag_ids = data.get('tag_ids', [])
        
        if not customer_ids or not tag_ids:
            return JsonResponse({'success': False, 'message': '请选择客户和标签'})
        
        for customer_id in customer_ids:
            for tag_id in tag_ids:
                CustomerTagRelation.objects.get_or_create(
                    customer_id=customer_id,
                    tag_id=tag_id
                )
        
        return JsonResponse({'success': True, 'message': f'已为 {len(customer_ids)} 个客户添加标签'})


@login_required
def tag_list_api(request):
    """获取所有标签（API）"""
    tags = CustomerTag.objects.all().values('id', 'name', 'color')
    return JsonResponse(list(tags), safe=False)



from django.db import IntegrityError

@login_required
def add_customer_manual(request):
    """手动添加客户"""
    if request.method == 'POST':
        company_name = request.POST.get('company_name')
        if not company_name:
            return JsonResponse({'success': False, 'message': '公司名称不能为空'})
        
        email = request.POST.get('email', '').strip()
        # 关键：如果 email 是空字符串，改为 None（数据库 NULL）
        if email == '':
            email = None
        
        # 获取选择的负责人
        assigned_sales_id = request.POST.get('assigned_sales')
        if not assigned_sales_id:
            return JsonResponse({'success': False, 'message': '请选择负责人'})
        
        try:
            customer = Customer.objects.create(
                company_name=company_name,
                contact_person=request.POST.get('contact_person', ''),
                email=email,
                phone=request.POST.get('phone', ''),
                country=request.POST.get('country', ''),
                level=request.POST.get('level', 'potential'),
                source='manual',
                notes=request.POST.get('notes', ''),
                is_deleted=False,
                assigned_sales_id=assigned_sales_id,  # 改这里：使用选择的负责人
            )
            
            # 设置部门（根据选择的负责人自动获取）
            from django.contrib.auth import get_user_model
            User = get_user_model()
            selected_sales = User.objects.filter(id=assigned_sales_id).first()
            if selected_sales and hasattr(selected_sales, 'profile') and selected_sales.profile.department:
                customer.department = selected_sales.profile.department
                customer.save()
            
            return JsonResponse({'success': True, 'customer_id': customer.id})
        except IntegrityError as e:
            if 'email' in str(e):
                return JsonResponse({'success': False, 'message': f'邮箱 "{email}" 已被其他客户使用，请更换邮箱'})
            else:
                return JsonResponse({'success': False, 'message': '客户信息重复，请检查公司名称或邮箱'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'添加失败：{str(e)}'})
    
    return JsonResponse({'success': False, 'message': '请求方法错误'})

@login_required
def update_order_cost(request, order_id):
    """更新订单成本/利润"""
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        
        # 获取成本明细
        fabric_cost = Decimal(request.POST.get('fabric_cost', '0') or '0')
        shipping_cost = Decimal(request.POST.get('shipping_cost', '0') or '0')
        customs_cost = Decimal(request.POST.get('customs_cost', '0') or '0')
        other_cost = Decimal(request.POST.get('other_cost', '0') or '0')
        
        total_cost = fabric_cost + shipping_cost + customs_cost + other_cost
        profit = order.subtotal - total_cost
        
        # 计算利润率
        if order.subtotal > 0:
            profit_margin = (profit / order.subtotal) * 100
        else:
            profit_margin = Decimal(0)
        
        # 保存
        order.total_cost = total_cost
        order.profit = profit
        order.profit_margin = profit_margin
        order.save()
        
        messages.success(request, '成本信息已更新')
        return redirect('order_detail', order_id=order.id)
    
    return redirect('order_detail', order_id=order_id)

# ==================== 数据统计大屏 ====================

from django.db.models import Sum, Count
from datetime import datetime, timedelta

def stats_dashboard(request):
    """数据统计大屏"""
    from .models import Order, Customer
    from send_logs.models import SendLog
    from django.db.models import Sum, Count, Q
    from django.contrib.auth import get_user_model
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    User = get_user_model()
    today = timezone.now().date()
    current_date = timezone.now().date()
    first_day_of_month = current_date.replace(day=1)
    
    # ========== 获取用户角色 ==========
    user_role = 'admin'
    user_dept = None
    user_sales = None
    if hasattr(request.user, 'profile'):
        user_role = request.user.profile.role
        user_dept = request.user.profile.department
        user_sales = request.user
    else:
        user_role = 'admin'
    
    # ========== 权限过滤函数 ==========
    def filter_by_role(queryset, model_type='order'):
        if user_role == 'sales':
            if model_type == 'order':
                return queryset.filter(customer__assigned_sales=user_sales)
            else:
                return queryset.filter(assigned_sales=user_sales)
        elif user_role == 'dept_leader' and user_dept:
            if model_type == 'order':
                return queryset.filter(customer__department=user_dept)
            else:
                return queryset.filter(department=user_dept)
        return queryset
    
    # ========== 订单数据（带权限） ==========
    all_orders = filter_by_role(Order.objects.all(), 'order')
    today_orders = filter_by_role(Order.objects.filter(order_date=current_date), 'order')
    month_orders = filter_by_role(Order.objects.filter(order_date__gte=first_day_of_month, order_date__lte=current_date), 'order')
    
    # 今日订单
    today_domestic = today_orders.filter(business_type='domestic')
    today_domestic_amount = today_domestic.aggregate(total=Sum('subtotal'))['total'] or 0
    today_domestic_count = today_domestic.count()
    today_international = today_orders.filter(business_type='international')
    today_international_amount = today_international.aggregate(total=Sum('subtotal'))['total'] or 0
    today_international_count = today_international.count()
    
    # 本月订单
    month_domestic = month_orders.filter(business_type='domestic')
    month_domestic_amount = month_domestic.aggregate(total=Sum('subtotal'))['total'] or 0
    month_domestic_count = month_domestic.count()
    month_international = month_orders.filter(business_type='international')
    month_international_amount = month_international.aggregate(total=Sum('subtotal'))['total'] or 0
    month_international_count = month_international.count()
    
    # 总销售额
    total_sales_domestic = filter_by_role(Order.objects.filter(business_type='domestic'), 'order').aggregate(total=Sum('subtotal'))['total'] or 0
    total_sales_international = filter_by_role(Order.objects.filter(business_type='international'), 'order').aggregate(total=Sum('subtotal'))['total'] or 0
    
    # ========== 月度统计（带权限） ==========
    months = []
    for i in range(11, -1, -1):
        month_date = today.replace(day=1) - timedelta(days=i*30)
        months.append(month_date.strftime('%Y-%m'))
    
    monthly_sales = []
    monthly_orders = []
    monthly_customers = []
    
    for month_str in months:
        year, month = map(int, month_str.split('-'))
        # 销售额
        sales = filter_by_role(Order.objects.filter(
            order_date__year=year, order_date__month=month,
            status__in=['confirmed', 'shipped', 'completed']
        ), 'order').aggregate(total=Sum('subtotal'))['total'] or 0
        monthly_sales.append(float(sales))
        # 订单数
        order_count = filter_by_role(Order.objects.filter(order_date__year=year, order_date__month=month), 'order').count()
        monthly_orders.append(order_count)
        # 新增客户数
        customer_count = filter_by_role(Customer.objects.filter(
            created_at__year=year, created_at__month=month, is_deleted=False
        ), 'customer').count()
        monthly_customers.append(customer_count)
    
    # ========== 客户数据（带权限） ==========
    all_customers = filter_by_role(Customer.objects.filter(is_deleted=False), 'customer')
    
    level_data = {
        'vip': all_customers.filter(level='vip').count(),
        'advanced': all_customers.filter(level='advanced').count(),
        'intermediate': all_customers.filter(level='intermediate').count(),
        'potential': all_customers.filter(level='potential').count(),
    }
    
    domestic_count = filter_by_role(Order.objects.filter(business_type='domestic'), 'order').count()
    international_count = filter_by_role(Order.objects.filter(business_type='international'), 'order').count()
    
    status_data = {
        'draft': filter_by_role(Order.objects.filter(status='draft'), 'order').count(),
        'confirmed': filter_by_role(Order.objects.filter(status='confirmed'), 'order').count(),
        'shipped': filter_by_role(Order.objects.filter(status='shipped'), 'order').count(),
        'completed': filter_by_role(Order.objects.filter(status='completed'), 'order').count(),
    }
    
    # ========== 产品销量排行（带权限） ==========
    product_sales = {}
    for order in all_orders:
        for item in order.items:
            product_name = item.get('product_name', '未知')
            amount = item.get('amount', 0)
            product_sales[product_name] = product_sales.get(product_name, 0) + amount
    
    product_top10 = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:10]
    product_names = [p[0] for p in product_top10]
    product_amounts = [p[1] for p in product_top10]
    
    # ========== 客户销售额排行（带权限） ==========
    customer_sales = {}
    for order in all_orders:
        customer_name = order.customer.company_name
        customer_sales[customer_name] = customer_sales.get(customer_name, 0) + order.subtotal
    
    customer_top10 = sorted(customer_sales.items(), key=lambda x: x[1], reverse=True)[:10]
    customer_names = [c[0] for c in customer_top10]
    customer_amounts = [c[1] for c in customer_top10]
    
    total_customers = all_customers.count()
    total_orders = all_orders.count()
    
    # 邮件回复率
    sent_count = SendLog.objects.count()
    reply_rate = 0
    
    # ========== 业绩排名（直接计算） ==========
    if user_role == 'sales':
        show_ranking = False
        ranking_scope = None
        sales_ranking = []
    elif user_role == 'dept_leader':
        show_ranking = True
        ranking_scope = 'department'
        sales_list = User.objects.filter(profile__role='sales', profile__department=user_dept)
        sales_data = []
        for sales in sales_list:
            customers = Customer.objects.filter(assigned_sales=sales, is_deleted=False)
            order_amount = Order.objects.filter(customer__in=customers).aggregate(total=Sum('subtotal'))['total'] or 0
            order_count = Order.objects.filter(customer__in=customers).count()
            customer_count = customers.count()
            sales_data.append({
                'id': sales.id,
                'username': sales.username,
                'order_amount': order_amount,
                'order_count': order_count,
                'customer_count': customer_count,
                'user_dept': sales.profile.department.name if hasattr(sales, 'profile') and sales.profile.department else '',
            })
        sales_ranking = sorted(sales_data, key=lambda x: x['order_amount'], reverse=True)[:10]
    else:
        show_ranking = True
        ranking_scope = 'all'
        sales_list = User.objects.filter(profile__role='sales')
        sales_data = []
        for sales in sales_list:
            customers = Customer.objects.filter(assigned_sales=sales, is_deleted=False)
            order_amount = Order.objects.filter(customer__in=customers).aggregate(total=Sum('subtotal'))['total'] or 0
            order_count = Order.objects.filter(customer__in=customers).count()
            customer_count = customers.count()
            sales_data.append({
                'id': sales.id,
                'username': sales.username,
                'order_amount': order_amount,
                'order_count': order_count,
                'customer_count': customer_count,
                'user_dept': sales.profile.department.name if hasattr(sales, 'profile') and sales.profile.department else '',
            })
        sales_ranking = sorted(sales_data, key=lambda x: x['order_amount'], reverse=True)[:10]    
    context = {
        'months': months,
        'monthly_sales': monthly_sales,
        'monthly_orders': monthly_orders,
        'monthly_customers': monthly_customers,
        'level_data': level_data,
        'domestic_count': domestic_count,
        'international_count': international_count,
        'status_data': status_data,
        'product_names': product_names,
        'product_amounts': product_amounts,
        'customer_names': customer_names,
        'customer_amounts': customer_amounts,
        'total_customers': total_customers,
        'total_orders': total_orders,
        'reply_rate': reply_rate,
        'total_sales_domestic': total_sales_domestic,
        'total_sales_international': total_sales_international,
        'today_domestic_amount': today_domestic_amount,
        'today_domestic_count': today_domestic_count,
        'today_international_amount': today_international_amount,
        'today_international_count': today_international_count,
        'month_domestic_amount': month_domestic_amount,
        'month_domestic_count': month_domestic_count,
        'month_international_amount': month_international_amount,
        'month_international_count': month_international_count,
        'show_ranking': show_ranking,
        'ranking_scope': ranking_scope,
        'sales_ranking': sales_ranking,
        'user_role': user_role,
        'user_dept_name': user_dept.name if user_dept else None,
    }
    
    return render(request, 'customers/stats_dashboard.html', context)
# ==================== 用户管理 ====================

from django.contrib.auth.hashers import make_password
from .models import User

@login_required
@login_required
def user_list(request):
    """用户列表"""
    from .models import UserProfile
    
    if not request.user.is_superuser and (not hasattr(request.user, 'profile') or request.user.profile.role != 'admin'):
        messages.error(request, '权限不足')
        return redirect('dashboard')
    
    users = User.objects.all().prefetch_related('profile', 'profile__department')
    return render(request, 'customers/user_list.html', {'users': users})

@login_required
def user_create(request):
    """创建用户"""
    from .models import UserProfile, Department
    from django.db import transaction
    
    # 权限检查
    if not request.user.is_superuser:
        if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != 'admin':
            return JsonResponse({'success': False, 'message': '权限不足'})
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')
        department_name = request.POST.get('department')
        email = request.POST.get('email', '')
        
        if not username or not password:
            return JsonResponse({'success': False, 'message': '用户名和密码不能为空'})
        
        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'message': '用户名已存在'})
        
        try:
            with transaction.atomic():
                # 创建用户
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    email=email,
                    is_staff=(role == 'admin'),
                )
                
                # 获取或创建部门
                department = None
                if department_name:
                    department, _ = Department.objects.get_or_create(name=department_name)
                
                # 创建或更新 Profile
                profile, created = UserProfile.objects.update_or_create(
                    user=user,
                    defaults={
                        'role': role,
                        'department': department
                    }
                )
                
                return JsonResponse({'success': True, 'message': '创建成功', 'department': department_name})
                
        except Exception as e:
            import traceback
            return JsonResponse({
                'success': False,
                'message': f'创建失败: {str(e)}',
                'trace': traceback.format_exc()
            }, status=500)
    
    return JsonResponse({'success': False, 'message': '仅支持POST请求'})

@login_required
def user_edit(request, user_id):
    """编辑用户"""
    from .models import UserProfile, Department
    
    if not request.user.is_superuser and (not hasattr(request.user, 'profile') or request.user.profile.role != 'admin'):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    user = get_object_or_404(User, id=user_id)
    
    # 确保用户有 profile
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        # 更新 Profile 表
        profile.role = request.POST.get('role', 'sales')
        
        dept_name = request.POST.get('department', '')
        if dept_name:
            try:
                department = Department.objects.get(name=dept_name)
                profile.department = department
            except Department.DoesNotExist:
                pass
        else:
            profile.department = None
        
        profile.save()
        
        # 更新 User 表
        user.email = request.POST.get('email', '')
        user.save()
        
        return JsonResponse({'success': True, 'message': '更新成功'})
    
    # GET 请求返回用户数据（用于编辑弹窗填充）
    data = {
        'success': True,
        'id': user.id,
        'username': user.username,
        'real_name': getattr(user, 'real_name', ''),
        'email': user.email,
        'phone': getattr(user, 'phone', ''),
        'role': profile.role if profile else 'sales',
        'department': profile.department.name if profile and profile.department else '',
    }
    return JsonResponse(data)

@login_required
def user_delete(request, user_id):
    """删除用户"""
    if not request.user.is_superuser and request.user.role != 'admin':
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    user = get_object_or_404(User, id=user_id)
    if user.id == request.user.id:
        return JsonResponse({'success': False, 'message': '不能删除自己'})
    
    user.delete()
    return JsonResponse({'success': True, 'message': '删除成功'})

from .models import CompanySetting

@login_required
def company_settings(request):
    """公司信息设置"""
    if not (request.user.is_superuser or request.user.role == 'admin'):
        messages.error(request, '您没有权限访问系统设置')
        return redirect('dashboard')
    
    # 获取或创建设置记录（只有一条）
    setting, created = CompanySetting.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        setting.company_name = request.POST.get('company_name', 'Raffinato CRM')
        setting.address = request.POST.get('address', '')
        setting.phone = request.POST.get('phone', '')
        setting.email = request.POST.get('email', '')
        setting.website = request.POST.get('website', '')
        setting.email_signature = request.POST.get('email_signature', '')
        
        if request.FILES.get('logo'):
            setting.logo = request.FILES['logo']
        
        setting.save()
        messages.success(request, '公司信息保存成功')
        return redirect('company_settings')
    
    return render(request, 'customers/company_settings.html', {'setting': setting})

# ==================== API：获取用户邮箱配置 ====================

@login_required
def api_user_email_configs(request):
    """获取当前用户的邮箱配置列表（用于前端选择）"""
    from .models import UserEmailConfig
    
    configs = UserEmailConfig.objects.filter(user=request.user, is_active=True)
    data = [{
        'id': c.id,
        'email': c.email,
        'from_name': c.from_name,
        'is_default': c.is_default,
    } for c in configs]
    return JsonResponse(data, safe=False)


# ==================== 沉睡客户配置 ====================

@login_required
def business_rules_settings(request):
    """业务规则设置"""
    if not (request.user.is_superuser or request.user.role == 'admin'):
        messages.error(request, '您没有权限访问系统设置')
        return redirect('dashboard')
    
    settings = BusinessRule.get_settings()
    
    if request.method == 'POST':
        settings.vip_downgrade_days = int(request.POST.get('vip_downgrade_days', 30))
        settings.advanced_downgrade_days = int(request.POST.get('advanced_downgrade_days', 60))
        settings.intermediate_downgrade_days = int(request.POST.get('intermediate_downgrade_days', 90))
        settings.dormant_days = int(request.POST.get('dormant_days', 180))
        settings.followup_reminder_days = int(request.POST.get('followup_reminder_days', 30))
        settings.daily_email_limit = int(request.POST.get('daily_email_limit', 200))
        settings.default_customer_level = request.POST.get('default_customer_level', 'potential')
        settings.save()
        
        messages.success(request, '业务规则保存成功')
        return redirect('business_rules_settings')
    
    # 等级选项
    level_choices = [
        ('potential', '潜在客户'),
        ('intermediate', '中级客户'),
        ('advanced', '高级客户'),
        ('vip', 'VIP客户'),
    ]
    
    return render(request, 'customers/business_rules_settings.html', {
        'settings': settings,
        'level_choices': level_choices,
    })

# ==================== 数据备份配置 ====================


@login_required
def backup_list(request):
    """备份列表页面"""
    if not (request.user.is_superuser or request.user.role == 'admin'):
        messages.error(request, '您没有权限访问')
        return redirect('dashboard')
    
    backups = BackupRecord.objects.all()
    return render(request, 'customers/backup_list.html', {'backups': backups})


@login_required
def backup_create(request):
    """创建备份"""
    if not (request.user.is_superuser or request.user.role == 'admin'):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    try:
        # 创建备份目录
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # 备份文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(backup_dir, f'crm_backup_{timestamp}.zip')
        
        # 数据库文件路径
        db_path = settings.BASE_DIR / 'db.sqlite3'
        
        # 创建 ZIP 备份
        with zipfile.ZipFile(backup_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # 备份数据库
            if os.path.exists(db_path):
                zipf.write(db_path, 'db.sqlite3')
            
            # 备份 media 文件（可选，如果太多可以注释）
            media_dir = settings.BASE_DIR / 'media'
            if os.path.exists(media_dir):
                for root, dirs, files in os.walk(media_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, settings.BASE_DIR)
                        zipf.write(file_path, arcname)
        
        # 计算文件大小
        file_size = os.path.getsize(backup_file) // 1024  # KB
        
        # 保存备份记录
        BackupRecord.objects.create(
            filename=os.path.basename(backup_file),
            file_size=file_size,
            backup_type='manual'
        )
        
        return JsonResponse({'success': True, 'message': '备份创建成功', 'filename': os.path.basename(backup_file)})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'备份失败：{str(e)}'})


@login_required
def backup_download(request, backup_id):
    """下载备份文件"""
    if not (request.user.is_superuser or request.user.role == 'admin'):
        messages.error(request, '您没有权限访问')
        return redirect('dashboard')
    
    backup = get_object_or_404(BackupRecord, id=backup_id)
    backup_path = os.path.join(settings.BASE_DIR, 'backups', backup.filename)
    
    if os.path.exists(backup_path):
        response = FileResponse(open(backup_path, 'rb'), as_attachment=True, filename=backup.filename)
        return response
    else:
        messages.error(request, '备份文件不存在')
        return redirect('backup_list')


@login_required
def backup_delete(request, backup_id):
    """删除备份文件"""
    if not (request.user.is_superuser or request.user.role == 'admin'):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    backup = get_object_or_404(BackupRecord, id=backup_id)
    backup_path = os.path.join(settings.BASE_DIR, 'backups', backup.filename)
    
    try:
        if os.path.exists(backup_path):
            os.remove(backup_path)
        backup.delete()
        return JsonResponse({'success': True, 'message': '删除成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'删除失败：{str(e)}'})


@login_required
def backup_restore(request):
    """恢复备份（上传备份文件）"""
    if not (request.user.is_superuser or request.user.role == 'admin'):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']
        
        # 检查文件类型
        if not backup_file.name.endswith('.zip'):
            return JsonResponse({'success': False, 'message': '请上传 ZIP 格式的备份文件'})
        
        try:
            # 保存上传的备份文件
            backup_dir = os.path.join(settings.BASE_DIR, 'backups')
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)
            
            file_path = os.path.join(backup_dir, backup_file.name)
            with open(file_path, 'wb') as f:
                for chunk in backup_file.chunks():
                    f.write(chunk)
            
            # 解压备份文件
            with zipfile.ZipFile(file_path, 'r') as zipf:
                # 解压到临时目录
                extract_dir = os.path.join(backup_dir, 'restore_temp')
                zipf.extractall(extract_dir)
                
                # 恢复数据库
                db_backup = os.path.join(extract_dir, 'db.sqlite3')
                if os.path.exists(db_backup):
                    import shutil
                    db_path = settings.BASE_DIR / 'db.sqlite3'
                    # 备份当前数据库
                    shutil.copy(db_path, os.path.join(backup_dir, f'db_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.sqlite3'))
                    # 恢复数据库
                    shutil.copy(db_backup, db_path)
                
                # 清理临时目录
                import shutil
                shutil.rmtree(extract_dir)
            
            # 记录恢复操作
            BackupRecord.objects.create(
                filename=f'restore_{backup_file.name}',
                file_size=backup_file.size // 1024,
                backup_type='manual'
            )
            
            return JsonResponse({'success': True, 'message': '备份恢复成功，请重启服务器'})
        
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'恢复失败：{str(e)}'})
    
    return JsonResponse({'success': False, 'message': '请选择备份文件'})


# ==================== 客户沟通记录配置 ====================

@login_required
def communication_list(request, customer_id):
    """获取客户的沟通记录列表（用于API）"""
    customer = get_object_or_404(Customer, id=customer_id, is_deleted=False)
    logs = CommunicationLog.objects.filter(customer=customer).order_by('-created_at')
    
    data = [{
        'id': log.id,
        'channel': log.channel,
        'channel_display': log.get_channel_display(),
        'direction': log.direction,
        'direction_display': log.get_direction_display(),
        'subject': log.subject,
        'content': log.content[:200] + '...' if len(log.content) > 200 else log.content,
        'content_full': log.content,
        'call_duration': log.call_duration,
        'call_duration_display': log.get_call_duration_display(),
        'attachment': log.attachment.url if log.attachment else None,
        'followup_needed': log.followup_needed,
        'followup_date': log.followup_date.strftime('%Y-%m-%d') if log.followup_date else None,
        'followup_note': log.followup_note,
        'created_by': log.created_by.real_name if log.created_by and log.created_by.real_name else (log.created_by.username if log.created_by else '系统'),
        'created_at': log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
    } for log in logs]
    
    return JsonResponse(data, safe=False)


from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
import json

@login_required
def communication_create(request, customer_id):
    """添加沟通记录"""
    if request.method == 'POST':
        try:
            customer = get_object_or_404(Customer, id=customer_id)
            
            # 处理通话时长
            call_duration = None
            call_minutes = request.POST.get('call_minutes')
            call_seconds = request.POST.get('call_seconds')
            if call_minutes and call_minutes != '0':
                call_duration = int(call_minutes or 0) * 60 + int(call_seconds or 0)
            
            # 处理跟进日期
            followup_date = None
            followup_date_str = request.POST.get('followup_date')
            if followup_date_str:
                from datetime import datetime
                followup_date = datetime.strptime(followup_date_str, '%Y-%m-%d').date()
            
            # 创建记录
            comm = CommunicationLog.objects.create(
                customer=customer,
                channel=request.POST.get('channel'),
                direction=request.POST.get('direction'),
                subject=request.POST.get('subject', ''),
                content=request.POST.get('content', ''),
                call_duration=call_duration,
                attachment=request.FILES.get('attachment'),
                followup_needed=request.POST.get('followup_needed') == 'on',
                followup_date=followup_date,
                followup_note=request.POST.get('followup_note', ''),
                created_by=request.user,
            )
            
            return JsonResponse({'success': True, 'message': '添加成功', 'id': comm.id})
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': '请求方法错误'})

@login_required
def communication_edit(request, log_id):
    """编辑沟通记录"""
    log = get_object_or_404(CommunicationLog, id=log_id)
    
    # 权限检查：只有记录人或管理员可以编辑
    if not (request.user.is_superuser or request.user.role == 'admin' or log.created_by == request.user):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    if request.method == 'POST':
        try:
            # 处理通话时长
            call_duration = None
            call_minutes = request.POST.get('call_minutes', '0')
            call_seconds = request.POST.get('call_seconds', '0')
            if call_minutes or call_seconds:
                call_duration = int(call_minutes or 0) * 60 + int(call_seconds or 0)
            
            log.channel = request.POST.get('channel')
            log.direction = request.POST.get('direction')
            log.subject = request.POST.get('subject', '')
            log.content = request.POST.get('content', '')
            log.call_duration = call_duration
            log.followup_needed = request.POST.get('followup_needed') == 'on'
            
            followup_date_str = request.POST.get('followup_date')
            if followup_date_str:
                from datetime import datetime
                log.followup_date = datetime.strptime(followup_date_str, '%Y-%m-%d').date()
            else:
                log.followup_date = None
            
            log.followup_note = request.POST.get('followup_note', '')
            
            # 处理附件
            if request.FILES.get('attachment'):
                log.attachment = request.FILES['attachment']
            
            log.save()
            
            return JsonResponse({'success': True, 'message': '更新成功'})
        
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'更新失败：{str(e)}'})
    
    return JsonResponse({'success': False, 'message': '请求方法错误'})


@login_required
def communication_delete(request, log_id):
    """删除沟通记录"""
    log = get_object_or_404(CommunicationLog, id=log_id)
    
    # 权限检查
    if not (request.user.is_superuser or request.user.role == 'admin' or log.created_by == request.user):
        return JsonResponse({'success': False, 'message': '权限不足'})
    
    log.delete()
    return JsonResponse({'success': True, 'message': '删除成功'})


@login_required
def communication_detail(request, log_id):
    """获取沟通记录详情（用于编辑弹窗）"""
    log = get_object_or_404(CommunicationLog, id=log_id)
    
    data = {
        'id': log.id,
        'channel': log.channel,
        'direction': log.direction,
        'subject': log.subject,
        'content': log.content,
        'call_duration': log.call_duration,
        'call_minutes': log.call_duration // 60 if log.call_duration else 0,
        'call_seconds': log.call_duration % 60 if log.call_duration else 0,
        'attachment': log.attachment.url if log.attachment else None,
        'followup_needed': log.followup_needed,
        'followup_date': log.followup_date.strftime('%Y-%m-%d') if log.followup_date else '',
        'followup_note': log.followup_note,
    }
    
    return JsonResponse(data)

@login_required
@csrf_exempt
@require_http_methods(['POST'])
def generate_ai_email(request, customer_id):
    """DeepSeek 生成邮件 - 支持风格和语言"""
    
    import json
    import requests
    from django.shortcuts import get_object_or_404
    from django.http import JsonResponse
    from .models import Customer
    
    customer = get_object_or_404(Customer, id=customer_id)
    
    subject_input = request.POST.get('subject', '')
    style = request.POST.get('style', 'formal')
    language = request.POST.get('language', 'chinese')
    
    if not subject_input:
        return JsonResponse({'success': False, 'message': '请输入邮件主题'})
    
    # 风格映射
    style_text = {
        'formal': '正式专业',
        'friendly': '友好亲切',
        'concise': '简洁明了'
    }
    style_cn = style_text.get(style, '正式专业')
    
    # 语言映射
    lang_name = '中文' if language == 'chinese' else 'English'
    
    api_key = 'sk-fb44df8bf89145b8ad7498b95936c699'
    
    # 根据语言选择提示词
    if language == 'chinese':
        prompt = f"""用户要求：
- 邮件主题必须是：{subject_input}
- 邮件风格：{style_cn}
- 客户公司：{customer.company_name}
- 联系人：{customer.contact_person or '先生/女士'}

请生成一封{style_cn}风格的邮件，直接返回JSON格式：
{{"subject": "{subject_input}", "body": "邮件正文"}}"""
    else:
        # 英文风格映射
        style_en = {
            'formal': 'formal and professional',
            'friendly': 'friendly and warm',
            'concise': 'concise and clear'
        }.get(style, 'formal and professional')
        
        prompt = f"""User requirements:
- Email subject must be: {subject_input}
- Email style: {style_en}
- Customer company: {customer.company_name}
- Contact person: {customer.contact_person or 'Sir/Madam'}

Please generate an email in {style_en} style, return directly in JSON format:
{{"subject": "{subject_input}", "body": "Email body"}}"""
    
    try:
        response = requests.post(
            "https://api.deepseek.com/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            },
            json={
                "model": "deepseek-chat",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.7,
                "max_tokens": 500
            },
            timeout=30
        )
        
        if response.status_code == 200:
            result = response.json()
            content = result['choices'][0]['message']['content']
            content = content.strip()
            if content.startswith('```json'):
                content = content[7:]
            if content.startswith('```'):
                content = content[3:]
            if content.endswith('```'):
                content = content[:-3]
            content = content.strip()
            
            data = json.loads(content)
            return JsonResponse({
                'success': True,
                'subject': data.get('subject', subject_input),
                'body': data.get('body', '')
            })
            
    except Exception as e:
        print(f"错误: {e}")
    
    return JsonResponse({
        'success': True,
        'subject': subject_input,
        'body': f"您好！\n\n关于：{subject_input}\n\n期待与您联系！\n\n祝好！"
    })
@login_required
@require_http_methods(['POST'])
def batch_import_results(request, task_id):
    """批量导入搜索结果"""
    from .models import SearchTask, SearchResult, Customer
    from .services.crawler_service import CrawlerService
    from .utils.extractors import extract_emails, extract_phones, extract_company_name
    
    # 移除 user=request.user 过滤
    task = get_object_or_404(SearchTask, id=task_id)
    
    # 可选：添加权限检查（如果任务有 created_by 字段）
    # if task.created_by != request.user:
    #     return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    
    results = task.results.filter(is_imported=False, is_ignored=False)
    
    imported_count = 0
    errors = []
    
    for result in results:
        try:
            crawler = CrawlerService()
            crawl_result = crawler.crawl(result.url)
            
            company_name = result.title
            emails = []
            phones = []
            
            if crawl_result['success']:
                html = crawl_result['content']
                emails = extract_emails(html)
                phones = extract_phones(html)
                company_name = extract_company_name(html, result.url) or result.title
            
            # 处理邮箱唯一约束
            email_value = emails[0] if emails else ''
            if not email_value:
                email_value = None
            else:
                if Customer.objects.filter(email=email_value).exists():
                    email_value = f"{email_value.split('@')[0]}+{result.id}@{email_value.split('@')[1]}"
            
            customer = Customer.objects.create(
                company_name=company_name or '待完善',
                email=email_value,
                phone=phones[0] if phones else '',
                created_at=timezone.now(),
            )
            
            result.company_name = company_name
            result.email = emails[0] if emails else ''
            result.phone = phones[0] if phones else ''
            result.is_imported = True
            result.imported_to = customer.id
            result.save()
            
            imported_count += 1
            
        except Exception as e:
            errors.append(f"{result.url}: {str(e)}")
    
    # 更新任务统计
    task.total_imported = task.results.filter(is_imported=True).count()
    task.save()
    
    return JsonResponse({
        'success': True,
        'imported_count': imported_count,
        'total': results.count(),
        'errors': errors
    })


@login_required
def export_search_results(request, task_id):
    """导出搜索结果到 Excel"""
    task = get_object_or_404(SearchTask, id=task_id)
    results = task.results.all()
    
    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "搜索结果"
    
    # 设置表头
    headers = ['序号', '公司名称', '网址', '标题', '邮箱', '电话', '联系人', '国家', '行业', '状态', '提取时间']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 填充数据
    for row, result in enumerate(results, 2):
        # 解析提取的数据
        extracted_data = result.extracted_data if result.extracted_data else {}
        
        ws.cell(row=row, column=1, value=row - 1)  # 序号
        ws.cell(row=row, column=2, value=result.company_name or '')
        ws.cell(row=row, column=3, value=result.url)
        ws.cell(row=row, column=4, value=result.title or '')
        ws.cell(row=row, column=5, value=result.email or '')
        ws.cell(row=row, column=6, value=result.phone or '')
        ws.cell(row=row, column=7, value=extracted_data.get('contact_person', ''))
        ws.cell(row=row, column=8, value=extracted_data.get('country', ''))
        ws.cell(row=row, column=9, value=extracted_data.get('industry', ''))
        
        # 状态
        if result.is_imported:
            status = '已导入'
        elif result.is_ignored:
            status = '已忽略'
        else:
            status = '待处理'
        ws.cell(row=row, column=10, value=status)
        
        ws.cell(row=row, column=11, value=result.created_at.strftime('%Y-%m-%d %H:%M') if result.created_at else '')
    
    # 调整列宽
    column_widths = [8, 25, 50, 40, 25, 15, 15, 12, 20, 10, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 20
    
    # 创建响应
    filename = f"search_results_{task.name}_{task.created_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def analyze_customer(request, customer_id):
    """AI 综合分析客户（双维度）"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        scoring_service = AIScoringService()
        result = scoring_service.analyze_customer(customer)
        
        return JsonResponse({
            'success': True,
            'ai_score': result['ai_score'],
            'ai_reasons': result['ai_reasons'],
            'order_score': result['order_score'],
            'order_reasons': result['order_reasons'],
            'order_stats': result['order_stats'],
            'tags': result['tags'],
            'customer_type': result['customer_type']['name'],
            'customer_type_icon': result['customer_type']['icon'],
            'customer_type_color': result['customer_type']['color'],
            'customer_type_strategy': result['customer_type']['strategy']
        })
    
    return JsonResponse({'success': False, 'message': '请使用POST请求'})

def order_export_pdf(request, order_id, doc_type):
    """导出订单 PDF 单据"""
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse, JsonResponse
    from .models import Order, CompanySetting
    from .services.pdf_generator import PDFGenerator
    
    order = get_object_or_404(Order, id=order_id)
    customer = order.customer
    
    company_setting = CompanySetting.objects.first()
    company_info = {
        'name': company_setting.company_name if company_setting else '',
        'address': company_setting.address if company_setting else '',
        'phone': company_setting.phone if company_setting else '',
        'email': company_setting.email if company_setting else '',
    }
    
    pdf_gen = PDFGenerator()
    
    if doc_type == 'quotation':
        buffer = pdf_gen.generate_quotation(order, customer, company_info)
        filename = f"Quotation_{order.order_no}.pdf"
    elif doc_type == 'proforma':
        buffer = pdf_gen.generate_invoice(order, customer, company_info, 'proforma')
        filename = f"Proforma_Invoice_{order.order_no}.pdf"
    elif doc_type == 'commercial':
        buffer = pdf_gen.generate_invoice(order, customer, company_info, 'commercial')
        filename = f"Commercial_Invoice_{order.order_no}.pdf"
    elif doc_type == 'packing':
        buffer = pdf_gen.generate_packing_list(order, customer, company_info)
        filename = f"Packing_List_{order.order_no}.pdf"
    else:
        return JsonResponse({'error': '无效的单据类型'}, status=400)
    
    # 获取 PDF 数据
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # 创建响应
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = len(pdf_data)
    
    return response

def track_open(request, log_id):
    """追踪邮件打开（1x1 透明像素）"""
    try:
        send_log = SendLog.objects.get(id=log_id)
        send_log.opened = True
        send_log.open_count += 1
        if not send_log.opened_at:
            send_log.opened_at = timezone.now()
        send_log.save()
        if send_log.customer:
            send_log.customer.last_contact_time = timezone.now()
            send_log.customer.save()
    except SendLog.DoesNotExist:
        pass
    
    pixel = b'GIF89a\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00!\xf9\x04\x01\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;'
    return HttpResponse(pixel, content_type='image/gif')

def track_click(request, log_id, link_id):
    """追踪链接点击"""
    original_url = request.GET.get('original_url', '/')
    try:
        send_log = SendLog.objects.get(id=log_id)
        
        # 更新点击统计
        send_log.clicked = True
        send_log.click_count += 1
        if not send_log.clicked_at:
            send_log.clicked_at = timezone.now()
        send_log.save()
        
        # 更新客户最后联系时间
        if send_log.customer:
            customer = send_log.customer
            customer.last_contact_time = timezone.now()
            customer.save()
    
    except SendLog.DoesNotExist:
        pass
    
    return redirect(original_url)

def check_reply(request):
    """检查邮件回复（通过 IMAP）"""
    # 这个功能需要配置邮箱 IMAP
    # 暂时返回 JSON，后续可扩展
    return JsonResponse({'status': 'pending', 'message': '功能开发中'})

# ==================== 邮件追踪功能 ====================
# 添加到 customers/views.py 文件末尾

import re
import hashlib
from django.core.mail import EmailMultiAlternatives
from send_logs.models import SendLog
from .models import UserEmailConfig

def process_links_for_tracking(content, log_id, request):
    """将邮件中的链接替换为追踪链接"""
    
    def replace_link(match):
        url = match.group(1)
        # 生成唯一链接 ID
        link_id = hashlib.md5(url.encode()).hexdigest()[:16]
        
        tracking_url = request.build_absolute_uri(
            reverse('track_click', args=[log_id, link_id])
        ) + f'?original_url={url}'
        
        return f'href="{tracking_url}"'
    
    # 替换所有 href 属性
    processed = re.sub(r'href="([^"]+)"', replace_link, content)
    
    return processed


@login_required
def send_tracking_email(request, customer_id):
    """发送带追踪的邮件"""
    from .models import Customer
    
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        subject = request.POST.get('subject')
        body = request.POST.get('body')
        
        # 创建发送记录
        send_log = SendLog.objects.create(
            customer=customer,
            subject=subject,
            content=body,
            sent_by=request.user,
            sent_at=timezone.now(),
            status='pending'
        )
        
        # 构建追踪像素
        tracking_url = request.build_absolute_uri(
            reverse('track_open', args=[send_log.id])
        )
        tracking_img = f'<img src="{tracking_url}" width="1" height="1" style="display:none;">'
        
        # 处理链接
        final_body = process_links_for_tracking(body, send_log.id, request)
        final_body += tracking_img
        
        # 获取发件邮箱配置
        email_config = UserEmailConfig.objects.filter(
            user=request.user,
            is_active=True
        ).first()
        
        if email_config:
            try:
                # 发送邮件
                msg = EmailMultiAlternatives(
                    subject=subject,
                    body=final_body,
                    from_email=email_config.email,
                    to=[customer.email],
                )
                msg.attach_alternative(final_body, "text/html")
                msg.send()
                
                send_log.status = 'sent'
                send_log.save()
                
                messages.success(request, '邮件发送成功！')
            except Exception as e:
                send_log.status = 'failed'
                send_log.error_message = str(e)
                send_log.save()
                messages.error(request, f'发送失败：{str(e)}')
        else:
            messages.error(request, '请先配置发件邮箱')
        
        return redirect('customer_detail', customer_id=customer.id)
    
    return JsonResponse({'error': '请使用POST请求'}, status=400)

@login_required
@require_http_methods(['POST'])
def batch_delete_search_tasks(request):
    """批量删除搜索任务"""
    from .models import SearchTask
    
    try:
        data = json.loads(request.body)
        task_ids = data.get('task_ids', [])
        
        if not task_ids:
            return JsonResponse({'success': False, 'message': '请选择要删除的任务'})
        
        tasks = SearchTask.objects.filter(id__in=task_ids)
        deleted_count = tasks.count()
        tasks.delete()
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'成功删除 {deleted_count} 个任务'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

def api_customer_search(request):
    """客户搜索 API（带权限过滤）"""
    q = request.GET.get('q', '')
    
    customers = Customer.objects.filter(is_deleted=False)
    
    # 权限过滤
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        role = request.user.profile.role
        if role == 'sales':
            customers = customers.filter(assigned_sales=request.user)
        elif role == 'dept_leader':
            dept = request.user.profile.department
            if dept:
                customers = customers.filter(department=dept)
        elif role == 'readonly':
            dept = request.user.profile.department
            if dept:
                customers = customers.filter(department=dept)
    
    if q:
        customers = customers.filter(company_name__icontains=q)
    
    customers = customers[:10]
    data = [{'id': c.id, 'company_name': c.company_name} for c in customers]
    return JsonResponse(data, safe=False)

@login_required
def get_template_content(request, customer_id):
    """获取模板内容"""
    from templates.models import EmailTemplate
    from customers.models import Customer
    
    template_id = request.GET.get('template_id')
    if not template_id:
        return JsonResponse({'success': False, 'message': '缺少模板ID'})
    
    try:
        template = EmailTemplate.objects.get(id=template_id)
        customer = Customer.objects.get(id=customer_id)
        
        # 替换模板中的变量
        subject = template.subject
        content = template.content
        
        # 替换客户相关变量
        subject = subject.replace('{company_name}', customer.company_name or '')
        subject = subject.replace('{contact_person}', customer.contact_person or '')
        subject = subject.replace('{country}', customer.country or '')
        
        content = content.replace('{company_name}', customer.company_name or '')
        content = content.replace('{contact_person}', customer.contact_person or '')
        content = content.replace('{country}', customer.country or '')
        
        # 替换公司相关变量（从 settings 或 CompanySetting 获取）
        from django.conf import settings
        my_name = getattr(settings, 'MY_NAME', '销售代表')
        my_company = getattr(settings, 'MY_COMPANY', '我的公司')
        
        subject = subject.replace('{my_name}', my_name)
        subject = subject.replace('{my_company}', my_company)
        content = content.replace('{my_name}', my_name)
        content = content.replace('{my_company}', my_company)
        
        return JsonResponse({
            'success': True,
            'subject': subject,
            'content': content
        })
        
    except EmailTemplate.DoesNotExist:
        return JsonResponse({'success': False, 'message': '模板不存在'})
    except Customer.DoesNotExist:
        return JsonResponse({'success': False, 'message': '客户不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def get_template_content_universal(request):
    """获取模板内容（不需要客户ID）"""
    from templates.models import EmailTemplate
    
    template_id = request.GET.get('template_id')
    if not template_id:
        return JsonResponse({'success': False, 'message': '缺少模板ID'})
    
    try:
        template = EmailTemplate.objects.get(id=template_id)
        
        # 批量发送时，先不替换客户变量，或者只替换通用变量
        subject = template.subject
        content = template.content
        
        return JsonResponse({
            'success': True,
            'subject': subject,
            'content': content
        })
        
    except EmailTemplate.DoesNotExist:
        return JsonResponse({'success': False, 'message': '模板不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

def search_task_list(request):
    """搜索任务列表"""
    from .models import SearchTask
    tasks = SearchTask.objects.all().order_by('-created_at')
    return render(request, 'customers/search_task_list.html', {'tasks': tasks})

# ========== 自动化邮件序列视图 ==========

from .models import EmailSequence, EmailSequenceStep, EmailSequenceQueue, CustomerSequenceState
from .models import Customer
from django.utils import timezone
from datetime import timedelta
from django.db import transaction

def email_sequence_list(request):
    """邮件序列列表"""
    sequences = EmailSequence.objects.all().order_by('-created_at')
    return render(request, 'customers/email_sequence_list.html', {'sequences': sequences})

def email_sequence_create(request):
    """创建邮件序列"""
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        trigger_type = request.POST.get('trigger_type')
        trigger_days = int(request.POST.get('trigger_days', 0))
        trigger_tag = request.POST.get('trigger_tag', '')
        
        sequence = EmailSequence.objects.create(
            name=name,
            description=description,
            trigger_type=trigger_type,
            trigger_days=trigger_days,
            trigger_tag=trigger_tag,
            status='draft'
        )
        messages.success(request, f'邮件序列 "{name}" 创建成功')
        return redirect('email_sequence_detail', pk=sequence.id)
    
    return render(request, 'customers/email_sequence_form.html')

def email_sequence_detail(request, pk):
    """邮件序列详情"""
    sequence = get_object_or_404(EmailSequence, id=pk)
    steps = sequence.steps.all().order_by('step_order')
    
    # 获取受影响的客户
    affected_customers = []
    if sequence.status == 'active':
        customers = Customer.objects.all()
        if sequence.trigger_tag:
            customers = customers.filter(tags__contains=sequence.trigger_tag)
        affected_customers = customers[:10]
    
    return render(request, 'customers/email_sequence_detail.html', {
        'sequence': sequence,
        'steps': steps,
        'affected_customers': affected_customers,
    })

def email_sequence_edit(request, pk):
    """编辑邮件序列"""
    sequence = get_object_or_404(EmailSequence, id=pk)
    
    if request.method == 'POST':
        sequence.name = request.POST.get('name')
        sequence.description = request.POST.get('description', '')
        sequence.trigger_type = request.POST.get('trigger_type')
        sequence.trigger_days = int(request.POST.get('trigger_days', 0))
        sequence.trigger_tag = request.POST.get('trigger_tag', '')
        sequence.save()
        messages.success(request, '邮件序列已更新')
        return redirect('email_sequence_detail', pk=sequence.id)
    
    return render(request, 'customers/email_sequence_form.html', {'sequence': sequence})

def email_sequence_delete(request, pk):
    """删除邮件序列"""
    sequence = get_object_or_404(EmailSequence, id=pk)
    name = sequence.name
    sequence.delete()
    messages.success(request, f'邮件序列 "{name}" 已删除')
    return redirect('email_sequence_list')

def email_sequence_toggle(request, pk):
    """启用/暂停/停止邮件序列"""
    sequence = get_object_or_404(EmailSequence, id=pk)
    action = request.GET.get('action', '')
    
    if action == 'activate':
        sequence.status = 'active'
        messages.success(request, f'邮件序列 "{sequence.name}" 已启用')
    elif action == 'pause':
        sequence.status = 'paused'
        messages.success(request, f'邮件序列 "{sequence.name}" 已暂停')
    elif action == 'stop':
        sequence.status = 'stopped'
        messages.success(request, f'邮件序列 "{sequence.name}" 已停止')
    
    sequence.save()
    return redirect('email_sequence_detail', pk=sequence.id)

def email_sequence_add_step(request, pk):
    """添加序列步骤"""
    sequence = get_object_or_404(EmailSequence, id=pk)
    
    if request.method == 'POST':
        step_order = EmailSequenceStep.objects.filter(sequence=sequence).count() + 1
        step = EmailSequenceStep.objects.create(
            sequence=sequence,
            step_order=step_order,
            wait_days=int(request.POST.get('wait_days', 0)),
            subject=request.POST.get('subject'),
            template=request.POST.get('template'),
            send_time=request.POST.get('send_time') or None,
            only_weekdays=request.POST.get('only_weekdays') == 'on',
        )
        messages.success(request, f'步骤 {step_order} 已添加')
        return redirect('email_sequence_detail', pk=sequence.id)
    
    return render(request, 'customers/email_sequence_step_form.html', {'sequence': sequence})

def email_sequence_edit_step(request, step_id):
    """编辑序列步骤"""
    step = get_object_or_404(EmailSequenceStep, id=step_id)
    sequence = step.sequence  # 明确获取 sequence
    
    if request.method == 'POST':
        step.wait_days = int(request.POST.get('wait_days', 0))
        step.subject = request.POST.get('subject')
        step.template = request.POST.get('template')
        step.send_time = request.POST.get('send_time') or None
        step.only_weekdays = request.POST.get('only_weekdays') == 'on'
        step.save()
        messages.success(request, '步骤已更新')
        return redirect('email_sequence_detail', pk=step.sequence.id)
    
    return render(request, 'customers/email_sequence_step_form.html', {
        'step': step,
        'sequence': sequence,  # 确保传递 sequence
    })

def email_sequence_delete_step(request, step_id):
    """删除序列步骤"""
    step = get_object_or_404(EmailSequenceStep, id=step_id)
    sequence_id = step.sequence.id
    step.delete()
    
    # 重新排序
    for idx, s in enumerate(EmailSequenceStep.objects.filter(sequence_id=sequence_id).order_by('step_order'), 1):
        s.step_order = idx
        s.save()
    
    messages.success(request, '步骤已删除')
    return redirect('email_sequence_detail', pk=sequence_id)

def email_sequence_queue(request):
    """邮件队列管理"""
    queue_items = EmailSequenceQueue.objects.filter(status='pending').order_by('scheduled_time')
    
    # 筛选
    status_filter = request.GET.get('status', '')
    if status_filter:
        queue_items = queue_items.filter(status=status_filter)
    
    return render(request, 'customers/email_sequence_queue.html', {
        'queue_items': queue_items,
        'status_filter': status_filter,
    })

def email_sequence_cancel(request, queue_id):
    """取消待发送邮件"""
    queue_item = get_object_or_404(EmailSequenceQueue, id=queue_id)
    queue_item.status = 'cancelled'
    queue_item.save()
    messages.success(request, '已取消发送')
    return redirect('email_sequence_queue')

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
import json

User = get_user_model()

def api_sales_list(request):
    """获取本部门的销售列表"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': '未登录'}, status=401)
    
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    # 直接查询所有用户，不通过 profile
    # 只要能登录的用户都显示
    all_users = User.objects.all()
    
    sales_list = [{
        'id': u.id,
        'username': u.username,
        'real_name': getattr(u, 'real_name', '')
    } for u in all_users]
    
    print(f"返回销售列表: {sales_list}")  # 调试用
    
    return JsonResponse({'sales': sales_list})

@csrf_exempt
def api_assign_customer(request):
    """分配客户给销售"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '方法不允许'})
    
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': '未登录'})
    
    if hasattr(request.user, 'profile'):
        role = request.user.profile.role
        if role not in ['admin', 'dept_leader']:
            return JsonResponse({'success': False, 'message': '无权限'})
    else:
        return JsonResponse({'success': False, 'message': '无权限'})
    
    try:
        data = json.loads(request.body)
        customer_id = data.get('customer_id')
        sales_id = data.get('sales_id')
        notes = data.get('notes', '')
        
        from .models import Customer, CustomerAssignment
        
        customer = Customer.objects.get(id=customer_id)
        sales_user = User.objects.get(id=sales_id)
        
        # 创建分配记录
        CustomerAssignment.objects.create(
            customer=customer,
            sales_user=sales_user,
            assigned_by=request.user,
            notes=notes,
            is_active=True
        )
        
        customer.assigned_sales = sales_user
        customer.save()
        
        return JsonResponse({'success': True, 'message': '分配成功'})
        
    except Customer.DoesNotExist:
        return JsonResponse({'success': False, 'message': '客户不存在'})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '销售不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

def api_product_search(request):
    """产品搜索 API - 只返回有订单的产品"""
    q = request.GET.get('q', '')
    products = set()
    
    # 从所有订单中提取产品名称
    for order in Order.objects.all():
        for item in order.items:
            if item.get('product_name'):
                pname = item.get('product_name')
                if pname and pname != 'None':
                    if not q or q.lower() in pname.lower():
                        products.add(pname)
    
    products = sorted(products)[:20]
    data = [{'name': p} for p in products]
    return JsonResponse(data, safe=False)

# 添加到 customers/models.py 或 views.py
def add_customer_to_sequence(customer, sequence_id):
    """将客户添加到邮件序列"""
    from .models import EmailSequence, EmailSequenceQueue, CustomerSequenceState
    
    sequence = EmailSequence.objects.get(id=sequence_id)
    
    # 检查是否已在序列中
    if CustomerSequenceState.objects.filter(
        customer=customer, sequence=sequence
    ).exists():
        return
    
    # 获取第一步
    first_step = sequence.steps.filter(step_order=1).first()
    if not first_step:
        return
    
    # 计算发送时间
    scheduled_time = timezone.now()
    if sequence.trigger_days > 0:
        scheduled_time += timedelta(days=sequence.trigger_days)
    
    # 加入队列
    EmailSequenceQueue.objects.create(
        sequence=sequence,
        step=first_step,
        customer=customer,
        scheduled_time=scheduled_time,
        status='pending'
    )
    
    # 记录状态
    CustomerSequenceState.objects.create(
        customer=customer,
        sequence=sequence,
        current_step=1,
        status='active',
        entered_at=timezone.now()
    )

# ========== 邮件追踪视图 ==========
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.shortcuts import redirect
from django.utils import timezone
from send_logs.models import SendLog

def track_open(request, log_id):
    """追踪邮件打开 - 返回1x1透明像素"""
    try:
        send_log = SendLog.objects.get(id=log_id)
        if not send_log.opened_at:
            send_log.opened_at = timezone.now()
            send_log.save()
            print(f"📧 邮件已打开: {send_log.subject} -> {send_log.recipient}")
    except Exception as e:
        print(f"追踪打开失败: {e}")
    
    # 1x1 透明 GIF 像素
    pixel = bytes([
        0x47, 0x49, 0x46, 0x38, 0x39, 0x61, 0x01, 0x00, 0x01, 0x00,
        0x80, 0x00, 0x00, 0xff, 0xff, 0xff, 0x00, 0x00, 0x00, 0x21,
        0xf9, 0x04, 0x01, 0x00, 0x00, 0x00, 0x00, 0x2c, 0x00, 0x00,
        0x00, 0x00, 0x01, 0x00, 0x01, 0x00, 0x00, 0x02, 0x02, 0x44,
        0x01, 0x00, 0x3b
    ])
    return HttpResponse(pixel, content_type='image/gif')


def track_click(request, log_id, link_id):
    """追踪邮件链接点击"""
    url = request.GET.get('url', '')
    
    try:
        send_log = SendLog.objects.get(id=log_id)
        send_log.click_count = (send_log.click_count or 0) + 1
        send_log.last_click_at = timezone.now()
        if not send_log.clicked_at:
            send_log.clicked_at = timezone.now()
        send_log.save()
        print(f"🔗 邮件链接被点击: {send_log.subject} -> {send_log.recipient} (点击次数: {send_log.click_count})")
    except Exception as e:
        print(f"追踪点击失败: {e}")
    
    if url:
        return redirect(url)
    return HttpResponseNotFound()


def track_stats(request, log_id):
    """查看邮件统计（API接口）"""
    try:
        send_log = SendLog.objects.get(id=log_id)
        return JsonResponse({
            'id': send_log.id,
            'subject': send_log.subject,
            'recipient': send_log.recipient,
            'status': send_log.status,
            'sent_at': send_log.sent_at,
            'opened_at': send_log.opened_at,
            'click_count': send_log.click_count,
            'clicked_at': send_log.clicked_at,
            'last_click_at': send_log.last_click_at,
        })
    except SendLog.DoesNotExist:
        return JsonResponse({'error': '记录不存在'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ========== 退订管理视图 ==========
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from send_logs.models import SendLog
from customers.models import UnsubscribeBlacklist, Customer, EmailSequenceQueue

def unsubscribe(request, log_id):
    """邮件退订页面"""
    send_log = get_object_or_404(SendLog, id=log_id)
    email = request.GET.get('email', send_log.recipient)
    
    # 检查是否已经退订
    already_unsubscribed = UnsubscribeBlacklist.objects.filter(email=email).exists()
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '用户主动退订')
        
        # 获取客户
        customer = Customer.objects.filter(email=email).first()
        
        # 添加到黑名单
        blacklist, created = UnsubscribeBlacklist.objects.get_or_create(
            email=email,
            defaults={
                'customer': customer,
                'unsubscribe_reason': reason,
                'unsubscribed_from': f'邮件 ID: {log_id}',
                'ip_address': request.META.get('REMOTE_ADDR', ''),
            }
        )
        
        if not created:
            # 已存在，更新原因
            blacklist.unsubscribe_reason = reason
            blacklist.unsubscribed_from = f'邮件 ID: {log_id}'
            blacklist.save()
        
        # 如果有关联客户，标记邮箱无效
        if customer:
            customer.email_invalid = True
            customer.email_bounced = True
            customer.bounce_reason = f'用户退订: {reason}'
            customer.save()
            
            # 取消该客户所有待发送的队列邮件
            EmailSequenceQueue.objects.filter(
                customer=customer,
                status='pending'
            ).update(status='cancelled', error_message=f'用户退订: {reason}')
        
        # 更新发送记录
        send_log.is_bounced = True
        send_log.bounce_type = 'unsubscribe'
        send_log.bounce_reason = reason
        send_log.save()
        
        return render(request, 'customers/unsubscribe_success.html', {
            'email': email,
            'already': False,
        })
    
    # GET 请求：显示退订确认页面
    return render(request, 'customers/unsubscribe_confirm.html', {
        'email': email,
        'log_id': log_id,
        'already_unsubscribed': already_unsubscribed,
    })


def unsubscribe_api(request):
    """API式退订（一键退订）"""
    email = request.GET.get('email')
    token = request.GET.get('token')
    
    if not email:
        return JsonResponse({'error': '缺少邮箱参数'}, status=400)
    
    # 简单的 token 验证（可选）
    # token 可以是 send_log.id 的加密
    customer = Customer.objects.filter(email=email).first()
    
    blacklist, created = UnsubscribeBlacklist.objects.get_or_create(
        email=email,
        defaults={
            'customer': customer,
            'unsubscribe_reason': 'API一键退订',
            'ip_address': request.META.get('REMOTE_ADDR', ''),
        }
    )
    
    if customer:
        customer.email_invalid = True
        customer.save()
        EmailSequenceQueue.objects.filter(customer=customer, status='pending').update(
            status='cancelled', error_message='用户退订'
        )
    
    return JsonResponse({'success': True, 'message': '已成功退订'})

# ========== 邮件统计报表视图 ==========
from django.db.models import Count, Q, Avg
from django.utils import timezone
from datetime import timedelta
from send_logs.models import SendLog

def email_stats_dashboard(request):
    """邮件统计仪表盘"""
    
    # 时间范围筛选
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    # 基础数据
    total_sent = SendLog.objects.filter(sent_at__gte=start_date).count()
    total_opened = SendLog.objects.filter(sent_at__gte=start_date, opened_at__isnull=False).count()
    total_clicked = SendLog.objects.filter(sent_at__gte=start_date, click_count__gt=0).count()
    total_bounced = SendLog.objects.filter(sent_at__gte=start_date, is_bounced=True).count()
    total_unsubscribed = SendLog.objects.filter(sent_at__gte=start_date, bounce_type='unsubscribe').count()
    
    # 计算比率
    open_rate = round((total_opened / total_sent * 100) if total_sent > 0 else 0, 1)
    click_rate = round((total_clicked / total_sent * 100) if total_sent > 0 else 0, 1)
    bounce_rate = round((total_bounced / total_sent * 100) if total_sent > 0 else 0, 1)
    
    # 每日趋势数据（用于图表）
    daily_stats = []
    for i in range(days, -1, -1):
        date = timezone.now() - timedelta(days=i)
        date_start = date.replace(hour=0, minute=0, second=0)
        date_end = date.replace(hour=23, minute=59, second=59)
        
        sent = SendLog.objects.filter(sent_at__date=date.date()).count()
        opened = SendLog.objects.filter(sent_at__date=date.date(), opened_at__isnull=False).count()
        clicked = SendLog.objects.filter(sent_at__date=date.date(), click_count__gt=0).count()
        
        daily_stats.append({
            'date': date.strftime('%m/%d'),
            'sent': sent,
            'opened': opened,
            'clicked': clicked,
        })
    
    # 按主题统计（热门邮件）
    top_subjects = SendLog.objects.filter(sent_at__gte=start_date).values('subject').annotate(
        sent=Count('id'),
        opened=Count('id', filter=Q(opened_at__isnull=False)),
        clicked=Count('id', filter=Q(click_count__gt=0)),
    ).order_by('-sent')[:10]
    
    # 按客户统计（活跃客户）
    top_customers = SendLog.objects.filter(sent_at__gte=start_date).values('customer__company_name').annotate(
        sent=Count('id'),
        opened=Count('id', filter=Q(opened_at__isnull=False)),
        clicked=Count('id', filter=Q(click_count__gt=0)),
    ).order_by('-opened')[:10]
    
    context = {
        'days': days,
        'total_sent': total_sent,
        'total_opened': total_opened,
        'total_clicked': total_clicked,
        'total_bounced': total_bounced,
        'total_unsubscribed': total_unsubscribed,
        'open_rate': open_rate,
        'click_rate': click_rate,
        'bounce_rate': bounce_rate,
        'daily_stats': daily_stats,
        'top_subjects': top_subjects,
        'top_customers': top_customers,
    }
    
    return render(request, 'customers/email_stats_dashboard.html', context)


def email_stats_detail(request):
    """邮件统计详情（数据表）"""
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    # 获取所有发送记录
    logs = SendLog.objects.filter(sent_at__gte=start_date).select_related('customer').order_by('-sent_at')
    
    # 搜索
    search = request.GET.get('search', '')
    if search:
        logs = logs.filter(
            Q(recipient__icontains=search) |
            Q(subject__icontains=search)
        )
    
    # 分页
    page = int(request.GET.get('page', 1))
    page_size = 50
    total = logs.count()
    total_pages = (total + page_size - 1) // page_size
    logs = logs[(page - 1) * page_size:page * page_size]
    
    context = {
        'logs': logs,
        'days': days,
        'search': search,
        'page': page,
        'total_pages': total_pages,
        'total': total,
    }
    
    return render(request, 'customers/email_stats_detail.html', context)


def email_stats_api(request):
    """邮件统计API（返回JSON）"""
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    stats = SendLog.objects.filter(sent_at__gte=start_date).aggregate(
        total_sent=Count('id'),
        total_opened=Count('id', filter=Q(opened_at__isnull=False)),
        total_clicked=Count('id', filter=Q(click_count__gt=0)),
        total_bounced=Count('id', filter=Q(is_bounced=True)),
    )
    
    return JsonResponse(stats)


